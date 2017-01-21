# -*- coding: utf-8 -*- 

import os
from LearnDriving.settings import TIME_ZONE
#from bs4 import BeautifulSoup
#import requests
from django.db.utils import IntegrityError
from math import ceil
SETTINGS_DIR = os.path.dirname(__file__)
MEDIA_ROOT = os.path.join(SETTINGS_DIR, 'media')
import django 
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "LearnDriving.settings")
django.setup()


#os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'LearnDriving.settings')
from django.db.models import Q,F
import xlrd,datetime
from django.core.exceptions import MultipleObjectsReturned
from unidecode import unidecode
import tempfile
import zipfile
import re
import random
#from pytz import timezone
from time import strftime, strptime
from django.utils import timezone
import pytz

from exceptions import AttributeError
from django.forms.fields import DateField
from django.utils.timezone import localtime
from django.http.response import HttpResponse
from itertools import chain
from django.utils.safestring import mark_safe
from django.db.models.expressions import Case, When
from django.db.models.fields import IntegerField

from rnoc.forms import D4_DATETIME_FORMAT, D4_DATE_ONLY_FORMAT
from rnoc.models import Tram, Mll, DoiTac, SuCo,\
    CaTruc, UserProfile, TrangThai, DuAn, ThaoTacLienQuan, ThietBi,\
    EditHistory, Lenh, FaultLibrary, NguyenNhan, Tinh, BSCRNC,\
    SiteType, BCNOSS, BTSType, UPE, QuanHuyen, DiaBan, TracNghiem, DapAn,\
    Component, PhuongXa, TanSo2G, Brand, MSC, CauHinh, Comment, TanSo3G, NhaTram


DATE_FORMAT_FOR_BCN = '%d/%m/%Y'
TIME_FORMAT_FOR_BCN =  '%H:%M'
DATETIME_FORMAT_FOR_BCN = '%d/%m/%Y %H:%M:%S'
def unique_list(seq):
    seen = set()
    seen_add = seen.add
    return [x for x in seq if not (x in seen or seen_add(x))]

def read_line(path,split_item):
    f =  open(path, "r")
    content = f.read().decode('utf-8')
    content = content.split(split_item)
    f.close()
    return content


def read_file_from_disk (path):
    f =  open(path, "rb") 
    a = f.read().decode('utf-8')
    f.close()
    return a


def save_file_to_disk(path,content, is_over_write):
    if is_over_write:
        with open(path, "wb") as f:
            f.write(content.encode('utf-8'))
    else:
        with open(path, "ab") as f:
            f.write(content.encode('utf-8'))
import pytz
'''
def convert_awaredate__time_to_local(d):
    est=pytz.timezone('US/Eastern')
    d.astimezone(est)
    return d
'''
def awaredate_time_to_local(d):
    d = localtime(d)
    return d
def local_a_naitive(d,timezone = TIME_ZONE ):#'Asia/Bangkok'
    eastern = pytz.timezone(timezone)
    loc_dt = eastern.localize(d)
    return loc_dt
def read_excel_cell(worksheet,row_index,curr_col):
    #print  'curr_col row_index ',curr_col, row_index
    cell_value = worksheet.cell_value(row_index, curr_col)
    #print 'cell_value',cell_value
    return cell_value

class Excel_2_3g(object):
    offset_begin_data_row = 0
    for_purpose_thong_bao = False
    check_type_for_BCN= False
    allow_create_one_instance_if_not_exit = True
    using_function_if_empty_fields_or_field_not_in_excel_field =[]# nhung cai field ma excel = rong van dung fucntion de gan gia tri cho field, vi du nhu field namekhong dau
    is_import_from_exported_file = False
    added_foreinkey_types = 0 # cai nay dung de tinh so luong du an, hoac thietbi, duoc add, neu nhieu qua thi stop
    max_length_added_foreinkey_types = 11500
    backwards_sequence =[]
    many2manyFields = []
    update_or_create_main_item = ''#Site_ID_3G
    worksheet_name = u'Sheet1'
    begin_row=0
    manual_mapping_dict = {}
    mapping_function_to_value_dict = {}
    auto_map = False
    model = Tram
    created_number =0
    update_number = 0
    just_create_map_field = False
    first_calculate_field = []
    def __init__(self,workbook=None,import_ghi_chu=None):
        
        self.import_ghi_chu = import_ghi_chu   #It's file name
        self.workbook = workbook
        self.worksheet = self.workbook.sheet_by_name(self.worksheet_name)
        
        self.num_rows = self.worksheet .nrows - 1
        self.num_cols = self.worksheet.ncols - 1
        self.excel_dict = self.read_excel_dict() # excel_dict la ten cua cac cot lay trong file excel ra
        #print 'excel_dict',self.excel_dict
        #return false
        self.base_fields = {}
        self.missing_fields =[]
        model_fieldnames = [f.name for f in self.model._meta.fields if f.name!='id' ]
        if self.many2manyFields: # self.ModelClass._meta.many_to_many
                for x in self.many2manyFields:
                    if x not in model_fieldnames:
                        model_fieldnames.append(x)
        if self.is_import_from_exported_file:
            for f in  self.model._meta.fields:
                if (f.verbose_name in self.excel_dict)  :
                    if f.name =='id':
                        continue
                    else:
                        #print '@@@f.name ',f.name 
                        self.base_fields[f.name] = self.excel_dict.get(f.verbose_name)
                elif f.name.replace("_",' ') in self.excel_dict:
                    self.base_fields[f.name] = self.excel_dict.get(f.name.replace("_",' ') )
                else:
                    self.missing_fields.append(f.name)
        else:
            for fname in model_fieldnames:
                is_missing_field = True
                fname_lower = fname.lower()
                if self.auto_map and (fname_lower in self.excel_dict):
                    self.base_fields[fname] = self.excel_dict[fname_lower]
                    is_missing_field = False
                if fname in self.manual_mapping_dict: #manual_mapping_dict la manual , do minh tu tao anh xa fieldname voi ten cot cua file excel
                    is_missing_field = False
                    fieldname_compare_with_fn_excel = self.manual_mapping_dict[fname]
                    if isinstance(fieldname_compare_with_fn_excel, int):
                        self.base_fields.update({fname:fieldname_compare_with_fn_excel})
                    else:
                        fieldname_in_excel =  unidecode(fieldname_compare_with_fn_excel).lower().replace(' ','_') # file name format
                        if fieldname_in_excel in self.excel_dict:
                            self.base_fields[fname]= self.excel_dict[fieldname_in_excel] #= so thu tu cua column chua field do, vi du 5
                        else: # thieu cot nay hoac da bi doi ten                        
                            raise ValueError('trong file excel thieu cot %s '%fieldname_in_excel)
                if is_missing_field == True:
                    self.missing_fields.append(fname)
                  
        #base_fields = excel field co trong model field + using_function_if_empty_fields_or_field_not_in_excel_field            
        for fname in self.using_function_if_empty_fields_or_field_not_in_excel_field:
            if fname in model_fieldnames and fname not in self.base_fields:
                self.base_fields[fname] = 'column not in excel'
        print 'base_fields dict',self.base_fields
        
        self.convert_basefield_to_list_of_tuple()
        
        if self.just_create_map_field:#for test
            return None
        
        self.loop_through_row_and_insertdb()
        
        
        
    def read_excel_dict(self): #{verbose_name_in_excel: col_index,verbose_name_in_excel2: col_index2}
        dict_attrName_columnNumber_excel_not_underscore = {}
        row_index = self.begin_row
        curr_col = 0
        while curr_col <= self.num_cols:
            atrrname = read_excel_cell(self.worksheet, row_index,curr_col)
            if not self.is_import_from_exported_file:
                atrrname = unidecode(atrrname).lower().replace(" ","_") #neu import trong 2g,3g binh thuong thi attr  co gach duoi
            else:
                atrrname = atrrname.replace("_"," ") # Neu import tu file export thi attr la verbose
                
            dict_attrName_columnNumber_excel_not_underscore[atrrname ]=   curr_col
            curr_col +=1
        return dict_attrName_columnNumber_excel_not_underscore
    def convert_basefield_to_list_of_tuple(self):
        #self.main_field_index_excel_column = self.base_fields.pop(self.update_or_create_main_item) #index of main fields
        self.main_dict = {}
        if isinstance(self.update_or_create_main_item, str):
            self.update_or_create_main_item = (self.update_or_create_main_item,)
        for i in self.update_or_create_main_item:
            self.main_dict.update({i:self.base_fields.pop(i)})
        
        
        
        if self.backwards_sequence:
            #can phai sap xep lai theo chuan [(),()] theo thu tu
            base_fields_lists = [x for x in self.base_fields.iterkeys()]
            for x in self.backwards_sequence:
                if x in base_fields_lists:
                    base_fields_lists.remove(x)
                    base_fields_lists.append(x)
            self.odering_base_columns_list_tuple = [(x, self.base_fields[x]) for x in base_fields_lists]
        else:
            self.odering_base_columns_list_tuple = [(k,v) for k, v in self.base_fields.iteritems()]
        #print 'self.odering_base_columns_list_tuple',self.odering_base_columns_list_tuple
    
    
    def loop_through_row_and_insertdb(self):
        row_index = getattr(self,'row_index',self.begin_row) + self.offset_begin_data_row
        self.tram_co_trong_3g_location_but_not_in_db3g = 0
        main_field_karg = {}
        self.has_created_tram_instance = False
        while row_index < self.num_rows:
            self.skip = False
            row_index += 1
            self.row_index = row_index
            print 'row_index',row_index
            first_calculate_field = getattr(self, 'first_calculate_field',None)
            self.temp_obj = {}
            if first_calculate_field:
                for field in  first_calculate_field:
                    value = self.create_value_for_one_field(row_index,field,self.base_fields[field]) 
                    self.temp_obj [field] = value
            if self.check_type_for_BCN:
                n_nhan  = read_excel_cell(self.worksheet,row_index, self.excel_dict['n.nhan'])
                thoi_gian_cb  = read_excel_cell(self.worksheet,row_index, self.excel_dict['thoi_gian_cb'])
                comment_vnp =  read_excel_cell(self.worksheet,row_index, self.excel_dict['vnp-ghi_chu'])
                if (thoi_gian_cb!='' and int(thoi_gian_cb) < 10) or n_nhan !='SITE_OOS' or comment_vnp =='':
                    print 'exit'
                    continue
                loai_ne = read_excel_cell(self.worksheet,row_index, self.excel_dict['loai_ne'])
                nha_cc = read_excel_cell(self.worksheet,row_index, self.excel_dict['nha_cc'])
                if loai_ne=='BSC':
                    continue
                if ((loai_ne=='CELL' or loai_ne=='NODEB' ) and nha_cc=='Ericsson'):
                    if (loai_ne=='NODEB'):
                        continue
                    self.type_excel = '3G'
                    ten_object  = read_excel_cell(self.worksheet,row_index, self.excel_dict['ten_ne'])
                    if ten_object[-6]=='K' and int (ten_object[-5] )>6:#loai bo U900 HCM
                        continue
                elif(loai_ne=='NODEB' and nha_cc=='Nokia'):
                    self.type_excel = 'NSM'
                elif(loai_ne=='NODEB' and nha_cc=='Alcatel'):
                    self.type_excel = 'ALU'
                elif(loai_ne=='BTS' and nha_cc=='Ericsson'):
                    self.type_excel = 'SRN'
                    loai_su_co  = read_excel_cell(self.worksheet,row_index, self.excel_dict['loai_su_co'])
                    if int(loai_su_co)==8:
                        continue
                else:
                    self.type_excel = '2G'
            for main_field in self.main_dict:
                value = read_excel_cell(self.worksheet,row_index, self.main_dict[main_field])
                to_value_function = self.get_function(main_field) # function for main field
                if to_value_function:
                    value = to_value_function(value)
            
        
                main_field_karg.update({main_field:value})
            if self.skip == True:
                continue
            print 'main_field_karg',main_field_karg
            filters = self.model.objects.filter(**main_field_karg)
            if filters: # co db_row nay roi, update thoi
                print '*update'
                self.created_or_update = 0
                for self.obj in filters:# loop va gan gia tri vao self.obj
                    if not self.for_purpose_thong_bao:
                        self.setattr_for_all_fields_for_one_obj(row_index)
                    self.update_number +=1
                    #print '@ so luong instance duoc update,',self.update_number
            else: #tao moi
                print 'tao moi***'
                if self.allow_create_one_instance_if_not_exit:
                    self.created_or_update = 1 
                    if not self.for_purpose_thong_bao:  
                        print '***** quan trong',main_field_karg
                        
                        if self.check_type_for_BCN:
                            filter_coi_trung_khong = BCNOSS.objects.filter(object=main_field_karg["object"],gio_mat__lt = main_field_karg["gio_mat"] + datetime.timedelta(minutes = 9),
                                       gio_mat__gt = main_field_karg["gio_mat"] - datetime.timedelta(minutes = 9))
                            if len(filter_coi_trung_khong) > 0:
                                print  '@@@@@@@@@@@@bo qua object khong tao moi',main_field_karg
                                #raise ValueError('dfasdfdfdf')
                                continue
                        self.obj = self.model(**main_field_karg)
                        self.setattr_for_all_fields_for_one_obj(row_index)
                    self.created_number +=1
                    #print '@ so luong instance duoc tao moi(new),',self.created_number
                else:
                    self.tram_co_trong_3g_location_but_not_in_db3g +=1
                    #print '@ so luong instance tram_co_trong_3g_location_but_not_in_db3g,',self.tram_co_trong_3g_location_but_not_in_db3g
                    continue
        self.thong_bao =  u'''số dòng được đọc %s,
        số dòng được cập nhập %s, 
        Số dòng được lưu vào database %s'''%(row_index-self.begin_row-self.offset_begin_data_row,self.update_number,self.created_number )
        print '@ so luong instance duoc update,',self.update_number
        print '@ so luong instance duoc tao moi(new),',self.created_number
        #print '@ so luong instance tram_co_trong_3g_location_but_not_in_db3g,',self.tram_co_trong_3g_location_but_not_in_db3g
    
    def create_value_for_one_field(self,row_index,field,column_number):
        
        if column_number !='column not in excel':
            value =  read_excel_cell(self.worksheet, row_index,column_number)
        else:
            value = None# value = "not in excel'
        if value=='' or value ==u'—':
            value = None#null
        if value != None and isinstance(value, unicode):
            if re.match('blank|null',value,re.IGNORECASE): 
                value = None#null
        # su dung ham khi value !=None
        if  value !=None or field in self.using_function_if_empty_fields_or_field_not_in_excel_field:
            to_value_function = self.get_function(field)
            if to_value_function:
                value = to_value_function(value)
        return value
    def setattr_for_all_fields_for_one_obj(self,row_index):
        for field_tuple in self.odering_base_columns_list_tuple:
            field = field_tuple[0]
            column_number = field_tuple[1]
            if field not in self.temp_obj:
                value = self.create_value_for_one_field(row_index,field,column_number)
            else:
                value =  self.temp_obj[field]   
            if value ==u'bỏ' or (value==None and field in self.many2manyFields):# tranh loi Nonetype gi gi do voi m2m
                pass
            else:
                setattr(self.obj, field, value) # save
        self.obj.save() 
        su_dung_obj_function = getattr(self, 'su_dung_obj_function',None)
        if su_dung_obj_function:
            su_dung_obj_function()       
    def get_function(self,field):
        if field in self.mapping_function_to_value_dict:
            func_name = self.mapping_function_to_value_dict[field]
            
            to_value_function = getattr(self, func_name)
            return to_value_function
        else:
            try:
                method_of_field_name = 'value_for_'+field
                print 'func_name',method_of_field_name
                to_value_function = getattr(self, method_of_field_name) #
                return to_value_function
            except: # Ko co ham nao thay doi gia tri value
                return None
    def value_for_gio_mat(self,value,DATETIME_FORMAT_BCN = None):
        #print 'gio_mat_@@@',value,type(value)
        d = local_a_naitive(datetime.datetime.strptime(value, DATETIME_FORMAT_BCN))
        return d
    def value_for_nha_san_xuat_2G(self,cell_value):
        return_value = self.value_for_Cabinet(cell_value,bts_type_name ='2G')
        return  return_value
    def value_for_active_2G(self,value):
        return True
    def value_for_active_3G(self,value):# ham cua excel_to_2g
        return True
    def value_for_active_4G(self,value):
        return True        
    def value_for_tinh(self,value):#chung
        if value:
            try:
                instance = Tinh.objects.get(ma_tinh = value)
            except:
                return None
                self.added_foreinkey_types +=1
                if self.added_foreinkey_types > self.max_length_added_foreinkey_types:
                    raise ValueError("so luong m2m field qua nhieu, kha nang la ban da chon thu tu field tuong ung voi excel column bi sai")
                instance = Tinh(ma_tinh = value)
                instance.save()
            self.obj
            return instance
        else:
            return None
    def value_for_dia_chi_3G_khong_dau(self,cell_value):
        if cell_value==None:
            return None
        return unidecode(cell_value)
    def value_for_dia_chi_2G_khong_dau(self,cell_value):
        if cell_value==None:
            return None
        return unidecode(cell_value)
            
    def value_for_is_tram_co_du_thong_tin_3g(self,value):
        return True
    def value_for_import_ghi_chu(self,value,prefix='3G: ',insert_index=0):
        import_ghi_chu_old = getattr(self.obj, 'import_ghi_chu',None)
        import_ghi_chu_old_s=[]
        if import_ghi_chu_old:
            import_ghi_chu_old_s = import_ghi_chu_old.split('\n')
        now_string = datetime.datetime.now().strftime(D4_DATETIME_FORMAT)
        string = prefix + u'import từ file ' + self.import_ghi_chu + u' vào lúc ' + now_string
        if prefix=='':
            raise ValueError('prefix must define in agrument')
        else:
            da_co_ghi_chu_cho_G_nay = False
            for count,moi_import in enumerate(import_ghi_chu_old_s):
                if re.match(prefix,moi_import):
                    insert_index = count
                    import_ghi_chu_old_s[insert_index] = string
                    da_co_ghi_chu_cho_G_nay = True
                    break
            if da_co_ghi_chu_cho_G_nay:
                import_ghi_chu_old_s[insert_index] = string
            else:
                import_ghi_chu_old_s.insert(insert_index,string)# gia su 4G phai nam o index 2 nhung no dang nam o 1, ma dang import_ghi_chu 2G
            #except IndexError:
                #import_ghi_chu_old_s.insert(insert_index,string)
        return_value = '\n'.join(import_ghi_chu_old_s)
        return return_value
    def value_for_Site_type(self,value):
        return SiteType.objects.get_or_create(Name = u'Site BTS/NodeB')[0]
    def value_for_dateField(self,cell_value):
        try:
            date = datetime.datetime(1899, 12, 30)
            get_ = datetime.timedelta(int(cell_value)) # delte du lieu datetime
            get_col2 = str(date + get_)[:10] # convert date to string theo dang nao do
            value = get_col2 # moi them vo
            return value
        except:
            return None
   
    def value_for_RNC(self,value,bsc_or_rnc = 'RNC'):#chinh
        if bsc_or_rnc == 'RNC':
            which_brand = '3G'
        elif bsc_or_rnc == 'BSC':
            which_brand = '2G'
        brand = getattr(self.obj, 'brand_' + which_brand)
        print '**brand in value_for_rnc',brand
        if value:
            try:
                instance = Tram.objects.filter(Site_Name_1__icontains = value)[0]
            except IndexError:
                if self.added_foreinkey_types  > self.max_length_added_foreinkey_types:
                    raise ValueError("so luong m2m field qua nhieu, kha nang la ban da chon thu tu field tuong ung voi excel column bi sai")
                instance = Tram(Site_Name_1 = value,\
                                Site_type = SiteType.objects.get(Name = u'Site 0 (RNC,BSC)'),\
                                ngay_gio_tao = timezone.now(),nguoi_tao = User.objects.get(username = u'rnoc2'),\
                                ly_do_sua = u'Được tạo ra từ import Database Trạm vào lúc %s '%(datetime.datetime.now().strftime(D4_DATETIME_FORMAT))
                                )
                instance.save()
                
            #sua od ay
            try:
                instance = BSCRNC.objects.filter(Name__icontains = value,brand=brand,bsc_or_rnc = bsc_or_rnc)[0]
            except IndexError:
                if self.added_foreinkey_types  > self.max_length_added_foreinkey_types:
                    raise ValueError("so luong m2m field qua nhieu, kha nang la ban da chon thu tu field tuong ung voi excel column bi sai")
                instance = BSCRNC(Name = value, 
                                  brand=brand,
                                  bsc_or_rnc = bsc_or_rnc,
                                  ngay_gio_tao = timezone.now(),\
                                  nguoi_tao = User.objects.get(username = 'rnoc2'),\
                                  ly_do_sua = u'Được tạo ra từ import Database Trạm vào lúc %s '%(datetime.datetime.now().strftime(D4_DATETIME_FORMAT)
                                                                                                  )
                                  )
                try:
                    instance.save()
                except IntegrityError:#dup problem thi sua brand lai
                    instance = BSCRNC.objects.filter(Name__icontains = value)[0]
                    instance.bsc_or_rnc = bsc_or_rnc
                    instance.brand = brand
                    instance.save()
            return instance
        else:
            return None
        
    def value_for_BSC_2G(self,value):
        print '***',value
        return self.value_for_RNC(value,bsc_or_rnc = 'BSC')
    def value_for_Foreinkey(self,value,**karg):
        Class_pass_to = karg.pop('Class_pass_to')
        is_tao = karg.pop('is_tao',True)
        is_limit_so_luong_tao_foreign = karg.pop('is_limit_so_luong_tao_foreign',True)
        if value == None or value =='':
            return None
        #rs = Class_pass_to.objects.get_or_create(Name = value)
        try:
            tan_so_2G_instance = Class_pass_to.objects.get(Name = value)
        except:
            if is_tao:
                print 'okkkkkkkkkkkkkkkkkkkk'
                tan_so_2G_instance = Class_pass_to(Name  = value,**karg)
                self.added_foreinkey_types +=1
                tan_so_2G_instance.save()
            else:
                raise ValueError("is_tao = false va ....")
        #tan_so_2G_instance = rs[0]
        #if rs[1]:
            #self.added_foreinkey_types +=1
        if (self.added_foreinkey_types  > self.max_length_added_foreinkey_types) and is_limit_so_luong_tao_foreign:
                    raise ValueError("so luong m2m field qua nhieu, kha nang la ban da chon thu tu field tuong ung voi excel column bi sai")
        return  tan_so_2G_instance    
    def value_for_tan_so_2G(self,value,Class_pass_to = TanSo2G):
        if value == None or value =='':
            return None
        rs = Class_pass_to.objects.get_or_create(Name = value)
        tan_so_2G_instance = rs[0]
        if rs[1]:
            self.added_foreinkey_types +=1
        if self.added_foreinkey_types  > self.max_length_added_foreinkey_types:
                    raise ValueError("so luong m2m field qua nhieu, kha nang la ban da chon thu tu field tuong ung voi excel column bi sai")
        return  tan_so_2G_instance
    def value_for_brand_2G(self,value):
        return self.value_for_tan_so_2G(value,Class_pass_to = Brand)  
    def value_for_brand_3G(self,value):
        return self.value_for_tan_so_2G(value,Class_pass_to = Brand)          
    def value_for_Name(self,value):
        if value:
            return value.rstrip().lstrip()
        else:
            return ''
    def value_for_common_datefield(self,cell_value):
        try:
            date = datetime.datetime(1899, 12, 30)
            get_ = datetime.timedelta(int(cell_value)) # delta du lieu datetime
            value = date + get_
            return value
        except:
            return None
    def value_for_common_datefield_exported_type(self,cell_value):
        cell_value = re.sub("$'", "", cell_value)
        d = datetime.datetime.strptime(cell_value, '%d/%m/%Y')
        return d
    def value_for_Cabinet(self,value,brand=None,bts_type_name = None):#chinh
        if bts_type_name ==None:
            bts_type_name = '3G'
        bts_type = BTSType.objects.get_or_create(Name = bts_type_name)[0]
        if brand==None:
            brand = getattr(self.obj, 'brand' + '_' + bts_type_name)
        else:
            brand = Brand.objects.get_or_create(Name = brand)[0]
        if value == None or value =='':
            return None
        if isinstance(value, float):
            value = u'%0.f'%value
   
        try:
            thietbi = ThietBi.objects.get(Name=value,bts_type = bts_type,brand=brand)
        except ThietBi.DoesNotExist:
            thietbi = ThietBi(Name=value)
            self.added_foreinkey_types +=1
            if self.added_foreinkey_types  > self.max_length_added_foreinkey_types:
                raise ValueError("so luong m2m field qua nhieu, kha nang la ban da chon thu tu field tuong ung voi excel column bi sai")
            thietbi.brand = brand
            thietbi.is_duoc_tao_truoc = True
            thietbi.nguoi_tao = User.objects.get(username="rnoc2")
            thietbi.bts_type = bts_type
            thietbi.ly_do_sua = u'Được tạo ra từ import database Trạm'
            try:
                thietbi.save()
            except IntegrityError:
                thietbi = ThietBi.objects.get(Name=value)
                self.added_foreinkey_types -=1
                thietbi.brand = brand
                thietbi.is_duoc_tao_truoc = True
                thietbi.nguoi_tao = User.objects.get(username="rnoc2")
                thietbi.bts_type = bts_type
                thietbi.ly_do_sua = u'Được tạo ra từ import database Trạm'
                thietbi.save()
        return thietbi
    def value_for_common_VLAN_ID (self,cell_value):
        value = int(cell_value)
        return value   
    
    def value_for_nguoi_tao (self,cell_value):#chinh
        if self.created_or_update ==1:
            if cell_value==None:
                user = User.objects.get (username = 'rnoc2')
                return user
            else:
                user = User.objects.get (username = cell_value)
                return user
        else:
            return u'bỏ'
    def value_for_nguoi_sua_cuoi_cung (self,cell_value):
        user = User.objects.get (username = cell_value)
        return user

    def value_for_ngay_gio_tao(self,cell_value,datetimeformat = D4_DATETIME_FORMAT):#chinh
        if self.created_or_update ==1:
            if cell_value:
                cell_value = re.sub("^'", "", cell_value)
                cell_value = re.sub('\s$', '', cell_value)
                d = datetime.datetime.strptime(str(cell_value), datetimeformat)

                return local_a_naitive(d)
            else:
                #now = datetime.datetime.now()
                now = timezone.now()
                return now
        else:
            return u'bỏ'
    def value_for_ngay_gio_sua(self,value):
        return timezone.now()
    '''
    def value_for_ngay_gio_sua(self,cell_value):
        cell_value = re.sub("^'", "", cell_value)
        cell_value = re.sub(' $', '', cell_value)
        d = datetime.datetime.strptime(str(cell_value), D4_DATETIME_FORMAT)
        return d
    '''
     
    def value_for_excel_export_boolean(self,cell_value):#boolean not null allowed
        #print '@@@@@@@@@cell_value',cell_value
        if cell_value ==u'✘' or cell_value==None:
            return False
        else:
            return True
    def value_for_Name_khong_dau(self,cell_value):
        field_co_dau = self.obj.Name
        if field_co_dau==None:
            return None
        return unidecode(field_co_dau)




    
class ExcelImportTrangThai (Excel_2_3g):
    is_import_from_exported_file = True
    using_function_if_empty_fields_or_field_not_in_excel_field = ['color_code','is_cap_nhap_gio_tot','nguoi_tao','ngay_gio_tao','is_duoc_tao_truoc','Name_khong_dau']
    #backwards_sequence =['Name_khong_dau']#de lay gia tri nha_san_xuat_2G truoc
    auto_map = False
    just_create_map_field = False
    update_or_create_main_item = 'Name'
    worksheet_name = u'Sheet 1'
    mapping_function_to_value_dict ={'is_cap_nhap_gio_tot':'value_for_excel_export_boolean','is_duoc_tao_truoc':'value_for_excel_export_boolean'}
    manual_mapping_dict = {}
    model = TrangThai
    
    def value_for_color_code(self,cell_value):#boolean
        if cell_value:
            return cell_value
        else:
            return "#%06x" % random.randint(0, 0xFFFFFF)
    '''
    def value_for_is_cap_nhap_gio_tot(self,cell_value):#boolean not null allowed
        #if cell_value ==u'✔':
            #return True
        if cell_value ==u'✘' or cell_value==None:
            return False
        else:
            return True
    
    def value_for_is_duoc_tao_truoc(self,cell_value):
        if self.obj.Name  ==u"Raise sự kiện":
            return True
        else:
            return False
    '''
class ExcelImportSuCo(ExcelImportTrangThai):
    model = SuCo
    
class ExcelImportComponent(ExcelImportTrangThai):
    model = Component
    def value_for_thiet_bi(self,cell_value):#boolean
        if cell_value:
            return ThietBi.objects.get(Name=cell_value.split('*')[0])
        else:
            return None
class ExcelImportNguyenNhan(ExcelImportTrangThai):
    model = NguyenNhan
    def value_for_color_code(self,cell_value):#boolean
        if cell_value:
            return cell_value
        else:
            return 'green'
class ExcelImportDoiTac (ExcelImportTrangThai):
    model = DoiTac
class ExcelImportDoiTac_ungcuu (Excel_2_3g):
    using_function_if_empty_fields_or_field_not_in_excel_field = ['Name_khong_dau','nguoi_tao','ngay_gio_tao']
    manual_mapping_dict = {'Name':u'HO_TEN','Don_vi' :u'TINH_TP1','So_dien_thoai':u'DIEN_THOAI','Thong_tin_khac':u'TINH_TP','email':u'EMAIL'}
    model = DoiTac
    worksheet_name = u'Sheet1'
    update_or_create_main_item = 'Name'
    def value_for_So_dien_thoai(self,value):
        if value:
            value = str(value)
            
            rs = re.subn('\.0$', '', value, 1)
            if rs[1]:
                value = rs[0]
                if value[0] !='0':
                    value = '0' + value
                return value
            else:
                return value
        else:
            return None
    def value_for_Thong_tin_khac(self,value):
        if value:
            try:
                tinh = Tinh.objects.get(ma_tinh=value)
                value = value + u', ' + tinh.dia_ban
                return value
            except Tinh.DoesNotExist:
                return value
        else:
            return None
class ImportTinh(Excel_2_3g):
    using_function_if_empty_fields_or_field_not_in_excel_field = ['Name_khong_dau']
    manual_mapping_dict = {'Name':u'Tên tỉnh','ma_tinh' :u'TINH_TP',}
    model = Tinh
    worksheet_name = u'Sheet1'
    update_or_create_main_item = 'Name'
DATETIME_FORMAT_BCN = '%d/%m/%Y %H:%M:%S'


    
class ImportBCN2G(Excel_2_3g):
    #row_index = 3843
    #first_calculate_field =''
    check_type_for_BCN= True
    type_excel = ''
    backwards_sequence = ['BTS_thiet_bi']
    thietbi_add_count = 0
    begin_row=12
    using_function_if_empty_fields_or_field_not_in_excel_field = ['BSC_or_RNC']
    manual_mapping_dict = {'object':u'Tên NE','gio_mat' :u'Thời gian sự cố','gio_tot':u'Thời gian CLR',\
                           'code_loi' :u'Loại sự cố','vnp_comment':u'VNP-Ghi chú',\
                           'gio_canh_bao_ac' :u'Thời gian cảnh báo AC','tong_thoi_gian':u'Thời gian CB','BTS_thiet_bi':u'Nhà CC' ,'BTS_Type':u'Loại NE',
                           'tinh':u'Tỉnh/TP'
                           }
    model = BCNOSS
    #worksheet_name = u'Sheet1'
    update_or_create_main_item = (u'object','gio_mat')
    def value_for_tinh(self,value):#ImportBCN2G
        tinh_ins = Tinh.objects.get(ma_tinh=value)
        return tinh_ins
    def value_for_tong_thoi_gian(self,value):
        if value:
            return int(value)
        else:
            return None
    
    
    def value_for_BTS_thiet_bi(self,value):
        if self.type_excel == 'SRN':
            value = 'SingleRan'
            brand = 'Ericsson'
        elif value =='Nokia':
            value = 'NSN'
            brand = 'NSN'
        else:
            #value = self.type_excel 
            brand = value
        value = value +  ' Cabinet'
        return_value = self.value_for_Cabinet(value,brand=brand,bts_type_name =self.obj.BTS_Type.Name)
        
        if self.type_excel == 'ALU' and self.has_created_tram_instance:
            self.tram_instance.Cabinet = return_value
            self.tram_instance.save()
            self.has_created_tram_instance = False
        return  return_value
    
    
    def value_for_BTS_Type(self,value):
        if self.type_excel == '3G' or self.type_excel == 'ALU'  or self.type_excel == 'NSM':
            value = BTSType.objects.get(Name = '3G')
            return value
        if value == 'BTS':
            value = BTSType.objects.get(Name = '2G')
            return value
    
    def value_for_object(self,value):#tra ve bsc rnc luon
        if self.type_excel == 'ALU':
            object_name = value
            site_name1_for_look  =re.sub('^3G_','',object_name)
            print 'site_name1_for_look',site_name1_for_look
            try:
                tram_alu = Tram.objects.filter(Site_Name_1 = site_name1_for_look)[0]
                bsc_or_rnc = tram_alu.RNC
            except: #IndexError,DoesNotExist:
                user = User.objects.get(username="rnoc2")
                now = timezone.now()
                try:
                    bsc_or_rnc = BSCRNC.objects.get(Name = 'UNKNOWRNC')
                except:
                    bsc_or_rnc = BSCRNC(Name = 'UNKNOWRNC',nguoi_tao=user,ngay_gio_tao=now,bsc_or_rnc="RNC")
                    bsc_or_rnc.save()
                self.tram_instance = Tram(Site_Name_1 = site_name1_for_look,Site_ID_3G = 'ALU_'+ object_name,RNC=bsc_or_rnc,
                     nguoi_tao=user,ngay_gio_tao=now,ly_do_sua=u'Được tạo ra từ import báo cáo ngày %s'%now.strftime(D4_DATETIME_FORMAT),Site_type = SiteType.objects.get(Name = 'Site BTS/NodeB'))
                self.has_created_tram_instance = True
            self.BSC_or_RNC = bsc_or_rnc
            return object_name
            #truong hop ALU handle xong luong rnc va nodeb
        elif self.type_excel == 'NSM':
            pattern = u'^(.*?) (.*?)$'#HCRNC23 3G_BCH065K7_HCM
            m = re.search(pattern, value)
            object_name = m.group(2) 
            bsc_rnc_name=  'HC'+m.group(1).replace('-','')
        elif self.type_excel == '3G':#3g eric
            pattern = u'^(.*?) (.*?)\d(_.*?)$'#HCRNC23 3G_BCH065K7_HCM
            print '****************ten NE',value
            m = re.search(pattern, value)
            object_name = m.group(2) +  m.group(3)
            bsc_rnc_name=  m.group(1)
        else:#2G
            if self.type_excel =='SRN':#HCBS249_HCM 2G_HMO011E_HCM
                pattern = u'^(.*?) (.*?)$'
                kq = re.findall(pattern, value)
            elif self.type_excel =='2G':
                pattern_lists =[u'^(.*?)\(.*?[: ](.*?)\)',u'^(.*?)\((.*?)\)$'] #u'^(.*?)\((.*?)\)$'  cho truong hop BSC_401H_KGG(LD-RACH-GIA_KGG)
                for pattern in pattern_lists:
                    kq = re.findall(pattern, value)
                    if kq:
                        break
            bsc_rnc_name = kq[0][0]
            object_name =  kq[0][1]
            object_name = re.sub('^2G_', '', object_name)
        try:
            bsc_instance = BSCRNC.objects.filter(Name__icontains = bsc_rnc_name)[0]
        except IndexError:
            #raise ValueError("chua co RNC nay %s" %bsc_rnc_name)
            if self.added_foreinkey_types  > self.max_length_added_foreinkey_types:
                raise ValueError("so luong m2m field qua nhieu, kha nang la ban da chon thu tu field tuong ung voi excel column bi sai")
            bsc_instance = BSCRNC(Name = bsc_rnc_name,\
                                  ngay_gio_tao = timezone.now(),\
                                  nguoi_tao = User.objects.get(username = 'rnoc2'),\
                                  ly_do_sua = u'được tạo ra khi import bao cao ngay %s'%self.type_excel
                                  )
            bsc_instance.save()
        self.BSC_or_RNC = bsc_instance
        return object_name
            
            
    def value_for_vnp_comment(self,value):
        #if value =='':
            #return '__skip__'
        m = re.search(u'^(.*?)- \d -', value)
        value  = m.group(1)
        return value
    def value_for_BSC_or_RNC(self,value):
        return self.BSC_or_RNC

    def value_for_gio_canh_bao_ac(self,value):
        if value:
            #print 'gio_canh bao ac_@@@',value,type(value)
            d = local_a_naitive(datetime.datetime.strptime(value, DATETIME_FORMAT_BCN))
            return d
        else:
            return None
        
    def value_for_gio_mat(self,value):
        d = local_a_naitive(datetime.datetime.strptime(value, DATETIME_FORMAT_BCN))
        
        
        return d
    def value_for_gio_tot(self,value):
        if value:
            #print 'gio_tot_@@@',value,type(value)
            d = local_a_naitive(datetime.datetime.strptime(value, DATETIME_FORMAT_BCN))
            return d
        else:
            return None
class ImportKeoDai(ImportBCN2G):
    backwards_sequence = []
    check_type_for_BCN= True
    type_excel = ''
    backwards_sequence = ['BTS_thiet_bi']
    thietbi_add_count = 0
    begin_row=0
    using_function_if_empty_fields_or_field_not_in_excel_field = ['code_loi']
    check_type_for_BCN = False
    manual_mapping_dict = {'object':u'Tên phần tử mất liên lạc',
                           'gio_mat' :u'Thời gian sự cố',
                          'BTS_thiet_bi':u'Nhà cung cấp' ,
                          'BTS_Type':u'Hệ thống',
                           'tinh':u'Tỉnh/TP',
                           'vnp_comment':u'Nguyên nhân, quá trình điều hành và phối hợp xử lý sự cố với các đơn vị',
                           'kien_nghi_de_xuat':u'Kiến nghị, đề xuất',
                           }
    model = BCNOSS
    #worksheet_name = u'Sheet1'
    update_or_create_main_item = (u'object','gio_mat')
    def value_for_BTS_Type(self,value):
        return Excel_2_3g.value_for_Foreinkey(self,value,Class_pass_to = BTSType)
    def value_for_object(self,value):#tra ve bsc rnc luon
        return value
    def value_for_code_loi(self,value):
        return 8
    def value_for_vnp_comment(self,value):
        return value
class ImportTinh_diaban(Excel_2_3g):
    
    using_function_if_empty_fields_or_field_not_in_excel_field = []
    manual_mapping_dict = {'Name':u'Khu vực','dia_ban' :u'Địa bàn','ghi_chu':u'Trực UCTT'}
    model = Tinh
    worksheet_name = u'Sheet3'
    update_or_create_main_item = 'Name'  
    def value_for_dia_ban(self,value):
        dia_ban_instance = DiaBan.objects.get(Name = value)
        return dia_ban_instance
class ImportRNC(Excel_2_3g):
    auto_map = True
    using_function_if_empty_fields_or_field_not_in_excel_field = ['nguoi_tao','ngay_gio_tao','import_ghi_chu']
    manual_mapping_dict = {'Name':u'RNCID'}
    model = BSCRNC
    worksheet_name = u'Thong ke NodeB-RNC'
    update_or_create_main_item = 'Name'
    def value_for_VI_TRI_RNC(self,value):
        if value:
            return Tinh.objects.get(ma_tinh = value)
        else:
            return None
    def value_for_import_ghi_chu(self,value):
        return_value = super(Excel_3G,self).value_for_import_ghi_chu(value,'RNC: ',insert_index=0)
        return return_value
class Import_RNC_Tram(Excel_2_3g):
    using_function_if_empty_fields_or_field_not_in_excel_field = ['nguoi_tao','ngay_gio_tao','Site_type','dia_chi_2G_khong_dau']
    manual_mapping_dict = {'Site_Name_1':u'Name','dia_chi_2G':u'DIA CHI','dia_chi_3G':u'DIA CHI','dia_chi_2G_khong_dau':u'DIA CHI','MSC':u'MSS'}
    model = Tram
    worksheet_name = u'Sheet 1'
    update_or_create_main_item = 'Site_Name_1'

    def value_for_Site_type(self,value):
        return SiteType.objects.get_or_create(Name = u'Site 0 (RNC,BSC)')[0]
    def value_for_MSC(self,value):
        return_value = self.value_for_Foreinkey(value,Class_pass_to = MSC)
        return return_value

class Import_BSCRNC(Excel_2_3g):
    using_function_if_empty_fields_or_field_not_in_excel_field = ['nguoi_tao','ngay_gio_tao']
    manual_mapping_dict = {'Name':u'Name','MSC':u'MSC','SGSN':u'SGSN'}
    model = BSCRNC
    worksheet_name = u'Sheet 1'
    update_or_create_main_item = 'Name'
    
    def value_for_MSC(self,value):
        return_value = self.value_for_Foreinkey(value,Class_pass_to = MSC,ngay_gio_tao = timezone.now(),nguoi_tao= User.objects.get(username="rnoc2"),msc_or_sgsn='MSC')
        return return_value
    def value_for_SGSN(self,value):
        return_value = self.value_for_Foreinkey(value,Class_pass_to = MSC,ngay_gio_tao = timezone.now(),nguoi_tao= User.objects.get(username="rnoc2"),msc_or_sgsn='SGSN')
        return return_value
class ExcelImportDuAn(ExcelImportTrangThai):
    model = DuAn
    def value_for_color_code(self,cell_value):#boolean
        if cell_value:
            return cell_value
        else:
            #return "#%06x" % random.randint(0, 0xFFFFFF)
            return 'organe'

class ExcelImportThietBi(ExcelImportTrangThai):
    model = ThietBi
    def value_for_color_code(self,cell_value):#boolean
        if cell_value:
            return cell_value
        else:
            return 'purple'
class ExcelImportFaultLibrary(ExcelImportTrangThai):
    model = FaultLibrary
class ExcelImportThaoTacLienQuan (ExcelImportTrangThai):
    model = ThaoTacLienQuan
class ExcelImportLenh (ExcelImportDuAn):
    using_function_if_empty_fields_or_field_not_in_excel_field = ['color_code','is_cap_nhap_gio_tot','nguoi_tao','ngay_gio_tao','is_duoc_tao_truoc','Name_khong_dau',]
    backwards_sequence= ['Name_khong_dau']
    update_or_create_main_item = 'command'
    model = Lenh
    def value_for_thiet_bi(self,value):
        if value:
            try:
                tb = ThietBi.objects.get(Name=value)
                return tb
            except ThietBi.DoesNotExist:
                return None
        else:
            return None
    '''    
    def value_for_Name(self,value):
        #print '@@@@@@@@i want see'
        if value==None:
            #print '@@@@@@@@i want see2'
            return ''
        else:
            return value
    '''
class Excel_3G(Excel_2_3g):
    offset_begin_data_row =  0
    using_function_if_empty_fields_or_field_not_in_excel_field = ['Site_ID_3G','nguoi_tao','ngay_gio_tao',
                                                                  'active_3G','Site_type','import_ghi_chu','brand_3G','tinh','Site_ID_2G']
    auto_map = True
    many2manyFields = ['du_an']
    just_create_map_field = False
    update_or_create_main_item = 'Site_Name_1'
    worksheet_name = u'Ericsson 3G'
    backwards_sequence =['is_co_U2100_rieng','du_an','UPE','Cabinet','BSC_2G']
    manual_mapping_dict = {'Project_Text':2,'du_an':2,'Cell_1_Site_remote':u'Cell 1 (carrier 1)', \
                    'Cell_2_Site_remote':u'Cell 2 (Carrier 1)', 'Cell_3_Site_remote':u'Cell 3 (Carrier 1)',\
                     'Cell_4_Site_remote':u'Cell 4 (Carrier 2)', 'Cell_5_Site_remote':u'Cell 5 (Carrier 2)', 'Cell_6_Site_remote':u'Cell 6 (Carrier 2)', \
                     'Cell_7_Site_remote':u'Cell 7 (remote/U900/3 carrier)', 'Cell_8_Site_remote':u'Cell 8 (remote/U900/3 carrier)', 'Cell_9_Site_remote':u'Cell 9 (remote/U900/3 carrier)',\
                     'Cell_K_U900_PSI':u'Cell K (U900 PSI)','tinh':u'Count Province',
                     #'Site_ID_3G':'Site ID',
                     }
    mapping_function_to_value_dict = {'Ngay_Phat_Song_2G':'value_for_dateField',\
                                      'Ngay_Phat_Song_3G':'value_for_dateField',\
                                      'IUB_VLAN_ID':'value_for_int_to_string','MUB_VLAN_ID':'value_for_int_to_string',\
                                      'Cell_1_Site_remote':u'value_error_but_equal_42', \
                    'Cell_2_Site_remote':u'value_error_but_equal_42', 'Cell_3_Site_remote':u'value_error_but_equal_42',\
                     'Cell_4_Site_remote':u'value_error_but_equal_42', 'Cell_5_Site_remote':u'value_error_but_equal_42', 'Cell_6_Site_remote':u'value_error_but_equal_42', \
                     'Cell_7_Site_remote':u'value_error_but_equal_42', 'Cell_8_Site_remote':u'value_error_but_equal_42', 'Cell_9_Site_remote':u'value_error_but_equal_42',\
                     'Cell_K_U900_PSI':u'value_error_but_equal_42',
                                    }
    
    def value_for_tinh(self, value):#Excel_4G
        value = self.obj.Site_Name_1[-3:]
        return super(Excel_3G,self).value_for_tinh(value)
    def value_for_brand_3G(self,value):
        value ='Ericsson'
        return super(Excel_3G, self).value_for_brand_2G(value)
    def value_for_Cabinet(self,value):
        return_value = super(Excel_3G, self).value_for_Cabinet(value, bts_type_name = '3G')
        return  return_value
    def value_for_UPE (self,value):
        if value:
            try:
                instance = UPE.objects.get(Name = value,tinh = self.obj.tinh )
            except:
                try:
                    instance = UPE(Name = value,tinh = self.obj.tinh)
                    instance.save()
                except:
                    instance = None
            return instance
        else:
            return None
        
    def value_for_import_ghi_chu(self,value,prefix='',insert_index=0):
        return_value = super(Excel_3G,self).value_for_import_ghi_chu(value,'3G: ',insert_index=0)
        return return_value
    def value_error_but_equal_42(self,value):
        if value ==42:
            return None
        else:
            return value
    def value_for_du_an(self,cell_value):
        if self.created_or_update == 1 :#create
            self.obj.save()
        try:
            du_an = DuAn.objects.get(Name=cell_value)
        except:
            du_an = DuAn(Name=cell_value)
            self.added_foreinkey_types +=1
            if self.added_foreinkey_types > self.max_length_added_foreinkey_types:
                raise ValueError("so luong m2m field qua nhieu, kha nang la ban da chon thu tu field tuong ung voi excel column bi sai")
            du_an.type_2G_or_3G = '3G'
            du_an.is_duoc_tao_truoc = True
            du_an.nguoi_tao = User.objects.get(username="rnoc2")
            du_an.ngay_gio_tao = timezone.now()
            du_an.save()
        self.obj.du_an.add(du_an)
        return None
    
    def value_for_BSC_2G(self,value):
        if len(value)==7:
            value = value + '_HCM'
            return super(Excel_3G,self).value_for_RNC(value,bsc_or_rnc = 'BSC')
    def value_for_Site_ID_3G(self,cell_value):
        ex = re.subn('_U900$','',cell_value)
        if ex[1]:# co phat hien chuoi _U900 o trong
            self.is_U900_hay_U2100 = 'U900'
            same_3g_has_site_name1 = ex[0]
            try:# co tram U2100 tuong ung voi tram U900 nay
                samesite_instance = Tram.objects.get(Site_ID_3G = 'ERI_3G_' + same_3g_has_site_name1)
                samesite_instance.is_co_U900_rieng = True
                samesite_instance.save()
                print '*********is_co_U2100_rieng'
                self.obj.is_co_U2100_rieng = True 
            except Tram.DoesNotExist:
                print '***** du me cu vo cai nay hoai la sao'
                self.obj.is_co_U2100_rieng = False#sao cu chay lenh nay
        else:# tram U2100
            self.is_U900_hay_U2100 = 'U2100'
            try:# neu co tram U900 rieng
                samesite_instance = Tram.objects.get(Site_ID_3G = 'ERI_3G_' + cell_value + '_U900')
                samesite_instance.is_co_U2100_rieng = True
                samesite_instance.save()
                self.obj.is_co_U900_rieng = True
            except Tram.DoesNotExist:# khogn co tram U900 nao ung voi tram U2100 nay
                self.obj.is_co_U900_rieng = False
        value = 'ERI_3G_' + cell_value
        return value
    
    def value_for_Site_ID_2G(self,value):#1
        if value ==None:
            return u'bỏ'
        else:
            self.obj.active_2G = True
            return 'SRN_2G_' + value
        
    def value_for_Site_Name_2(self,cell_value):
        value = cell_value.replace('3G_','')
        return value
    def value_for_int_to_string (self,cell_value):
        value = int(cell_value)
        return value
    def value_for_Site_Name_1 (self,value):
        value = value.replace("3G_","")
        return value
    '''
    def value_for_is_co_U2100_rieng(self,value):
        print '************2'
        if self.is_U900_hay_U2100 == 'U900':
            self.obj.is_co_U2100_rieng = self.is_co_U2100_rieng
        elif self.is_U900_hay_U2100 == 'U2100':
            self.obj.is_co_U900_rieng = self.is_co_U900_rieng
        return None
    '''
    
D4_DATE_ONLY_FORMAT_gachngang = '%d-%m-%Y'    
class Excel_to_2g (Excel_2_3g):
    offset_begin_data_row =  0
    import_ghi_chu_text = 'Ericsson_Database_Ver_160'
    using_function_if_empty_fields_or_field_not_in_excel_field = ['nguoi_tao','active_2G','Site_ID_2G_Number','Site_type','import_ghi_chu',]
    backwards_sequence =['nha_san_xuat_2G','Site_ID_2G','Site_ID_2G_Number','quan_huyen','BSC_2G']
    auto_map = False
    just_create_map_field = False
    update_or_create_main_item = 'Site_Name_1'
    worksheet_name = u'Database 2G'
    mapping_function_to_value_dict = {#'Ngay_Phat_Song_2G':'value_for_dateField'
                                      
                                      }
    manual_mapping_dict = {'Ngay_Phat_Song_2G': u'Phát sóng','Site_Name_1':u'Tên BTS',
                           'dia_chi_2G':u'Địa chỉ',
                            'BSC_2G':u'Tên BSC',\
                    'LAC_2G':u'LAC',
                     'Nha_Tram':u'Nhà trạm', 'Ma_Tram_DHTT':u'Mã trạm ĐHTT', 'Cell_ID_2G':u'CellId', \
                    'cau_hinh_2G':u'Cấu hình',
                    'brand_2G':u'Nhà SX', 
                     'nha_san_xuat_2G':u'Nhà SX', 
                    'Site_ID_2G':u'Tên BTS',\
                    
                    'Long_2G':u'Tọa độ - Kinh độ','Lat_2G':u'Tọa độ - Vĩ độ',
                    'quan_huyen':u'Quận/Huyện',
                    'tinh':u'Mã tỉnh',
                    'dia_chi_2G_khong_dau':u'Địa chỉ'}
    def value_for_cau_hinh_2G(self,value):
        if isinstance(value, float):
            value = u'%0.f'%value
        value = value.replace(' ','/')
        if '/' not in value:
            rss = re.findall('\d', value)
            value ='/'.join(rss)
        
        return self.value_for_Foreinkey(value,Class_pass_to = CauHinh)
    def value_for_brand_2G(self,value):
        value = value.capitalize()
        if value =='Singleran':
            value = 'Ericsson' 
        return super(Excel_to_2g, self).value_for_brand_2G(value)
    def value_for_nha_san_xuat_2G(self,value):
        value = value + ' Cabinet'
        return_value = self.value_for_Cabinet(value,bts_type_name ='2G')
        return  return_value
    def value_for_quan_huyen(self,value):
        if value:
            try:
                instance = QuanHuyen.objects.get(Name = value,tinh = self.obj.tinh )
            except:
                instance = QuanHuyen(Name = value,tinh = self.obj.tinh)
                instance.save()
            return instance
        else:
            return None
    
    def value_for_import_ghi_chu(self,value):
        return_value = super(Excel_to_2g,self).value_for_import_ghi_chu(value,'2G: ',insert_index=1)
        return return_value
    def value_for_Ngay_Phat_Song_2G(self,value):
        rs = datetime.datetime.strptime(value, D4_DATE_ONLY_FORMAT_gachngang)
        return rs
    def value_for_Site_ID_2G_Number(self,value):#2
        value = self.obj.Cell_ID_2G
        if value:
            value = int(value)
            if value < 1000:
                value = str(value)[-2:]
            else:
                value = str(value)[-3:-1]
            return value
        else:
            return None
    
    def value_for_Site_Name_1 (self,cell_value):
        value = cell_value.replace("2G_","")
        return value
    def value_for_Site_ID_2G(self,cell_value):#3
        cell_value = re.sub('^2G_','',cell_value)
        '''
        if cell_value.startswith('2G_'):
            return None  # return none for not save to database this field
        else:
        '''
        prifix = self.obj.nha_san_xuat_2G.Name[0:3]
        if re.match('^sin',prifix,re.IGNORECASE):
            prifix = 'SRN'
            return u'bỏ'
        cell_value = prifix.upper() + '_2G_' + cell_value
        return  cell_value
    
    
### ngay 05/10/2015    
class Excel_2g_rnas(Excel_2_3g):#bo sung the 2g, quan huyen
    #for_purpose_thong_bao = False
    begin_row=1
    offset_begin_data_row = 0
    import_ghi_chu_text = 'file New BTS_2009_rnas.xlsx'
    using_function_if_empty_fields_or_field_not_in_excel_field = ['Site_type','import_ghi_chu','active_2G','nguoi_tao',
                                                                  'ngay_gio_tao']
    backwards_sequence =['Site_ID_2G_Number','quan_huyen','phuong_xa','BSC_2G','nha_san_xuat_2G','Site_ID_2G']
    auto_map = False
    just_create_map_field = False
    update_or_create_main_item = 'Site_Name_1'
    worksheet_name = u'BIM INFO'
    mapping_function_to_value_dict = {#'Ngay_Phat_Song_2G':'value_for_dateField'
                                      }

    manual_mapping_dict = {'Site_Name_1':u'Tên trên hệ thống',
                           'Site_ID_2G':u'Tên cho quản lý',
                           'brand_2G':u'Thiết bị',
                           'nha_san_xuat_2G':u'Thiết bị',
                           'Ngay_Phat_Song_2G':u'Ngày hoạt động',\
                           'quan_huyen':u'Quận/Huyện','tinh':u'Tỉnh/TP','phuong_xa':u'Phường/Xã',
                           'dia_chi_2G':u'Địa chỉ lắp đặt','dia_chi_2G_khong_dau':u'Địa chỉ lắp đặt',
                           'Long_2G':u'Longitude','Lat_2G':u'Latitude',
                           'BSC_2G':u'Tên BSC/RNC quản lý',
                           'cau_hinh_2G':u'Cấu hình',
                           'macro_or_ibs':u'Loại trạm'
                }
    
    def value_for_cau_hinh_2G(self,value):
        if isinstance(value, float):
            print value
            value = u'%0.f'%value
        value = value.replace(' ','/')
        if '/' not in value:
            rss = re.findall('\d', value)
            value ='/'.join(rss)
        
        return self.value_for_Foreinkey(value,Class_pass_to = CauHinh)
    def value_for_brand_2G(self,value):
        value = value.capitalize()
        return super(Excel_2g_rnas, self).value_for_brand_2G(value)
    
    def value_for_nha_san_xuat_2G(self,value):# chua biet tu gi
        
        if value == 'Ericsson': 
            value ='SingleRan'
        value = value.capitalize()
        value = value + ' Cabinet'
        return_value = self.value_for_Cabinet(value,bts_type_name ='2G')
        return  return_value
    def value_for_tinh(self, value):#Excel_4G
        value = self.obj.Site_Name_1[-3:]
        return super(Excel_2g_rnas,self).value_for_tinh(value)
    def value_for_Site_Name_1 (self,cell_value):
        value = re.sub('^2G_','',cell_value)
        print value
        return value
    def value_for_Site_ID_2G(self,cell_value):#4
        if self.obj.nha_san_xuat_2G !=None:
            prefix = self.obj.nha_san_xuat_2G.Name[0:3]
            if prefix =='Sin':
                prefix = 'SRN'
                return u'bỏ'
        else:
            prefix = 'UNK'
        cell_value = prefix.upper() + '_2G_' + self.obj.Site_Name_1
        return  cell_value
    def value_for_quan_huyen(self,value):
        if value:
            try:
                instance = QuanHuyen.objects.get(Name = value,tinh = self.obj.tinh )
            except:
                instance = QuanHuyen(Name = value,tinh = self.obj.tinh)
                instance.save()
            return instance
        else:
            return None
    def value_for_phuong_xa(self,value):
        if value:
            #if re.search('[a-zA-Z]+', value):
            if isinstance(value, unicode):
                value = re.sub(u'^xã ', u'Xã ', value, 0, re.IGNORECASE)
            else:
                return None
            try:
                instance = PhuongXa.objects.get(Name = value,tinh = self.obj.tinh,quan_huyen = self.obj.quan_huyen )
            except:
                instance = PhuongXa(Name = value,tinh = self.obj.tinh,quan_huyen = self.obj.quan_huyen)
                instance.save()
            return instance
        else:
            return None
    def value_for_import_ghi_chu(self,value):
        return_value = super(Excel_2g_rnas,self).value_for_import_ghi_chu(value,'2G: ',insert_index=1)
        return return_value
    def value_for_Ngay_Phat_Song_2G(self,value):
        rs = datetime.datetime.strptime(value, '%d/%m/%Y')
        return rs
class Excel_moto_onair(Excel_2_3g):# them duoc band, tu hm2,CI
    #for_purpose_thong_bao = False
    begin_row=0
    offset_begin_data_row = 0
    import_ghi_chu_text = 'CONFIG - ONAIR 2G MOTO W40.xlsx'
    using_function_if_empty_fields_or_field_not_in_excel_field = ['Site_ID_2G_Number','Site_type','import_ghi_chu','nguoi_tao','active_2G']
    backwards_sequence =['Cell_ID_2G','nha_san_xuat_2G','BSC_2G']#nha_san_xuat_2G sau brand
    auto_map = False
    just_create_map_field = False
    update_or_create_main_item = ['Site_Name_1']
    worksheet_name = u'Config_2G_W40'
    mapping_function_to_value_dict = {
                                      }
    manual_mapping_dict = {'Site_Name_1':u'Cell name','Site_ID_2G':u'BTS Name','nha_san_xuat_2G':u'CAB',\
                           'BSC_2G':'MBSC','Cell_ID_2G':'CI','tan_so_2G':'Freq Band','brand_2G':u'Vendor'
                }
    def value_for_brand_2G(self,value):
        value ='Motorola'
        return super(Excel_moto_onair, self).value_for_brand_2G(value)
    
    def value_for_Site_Name_1 (self,cell_value):
        print '**cell_value',cell_value
        rs = re.findall('(\d+):(.*?)$',cell_value)
        value = rs[0][1]
        self.Site_ID_2G_Number =  rs[0][0]
        print '******value',value,'value[-5]',value[-5]
        self.cell_nth =  int(value[-5])
        if self.cell_nth ==0:
            self.cell_nth =1
        value = value[:-5] + value[-4:]
        print value
        self.Site_Name_1 = value
        return value
    def value_for_Site_ID_2G(self,cell_value):#5
        prefix='MOT'
        cell_value = prefix.upper() + '_2G_' + self.Site_Name_1
        return  cell_value
    def value_for_Site_ID_2G_Number(self,cell_value):
        return  self.Site_ID_2G_Number
    
    def value_for_Cell_ID_2G(self,value):
        value = u'%0.f'%value
        print 'self.cell_nth',self.cell_nth
        cell_index = self.cell_nth - 1
        site2gnumber_old_text= getattr(self.obj,'Cell_ID_2G',None)
        if site2gnumber_old_text:
            site2gnumber_olds = site2gnumber_old_text.split(',')
        else:
            site2gnumber_olds = []
        for i in range(0,self.cell_nth):#'None,cell2,None'
            value_i = value if i ==cell_index else 'None'
            try:
                site2gnumber_olds[i]
            except IndexError:
                site2gnumber_olds.append(value_i)
        print 'site2gnumber_olds',site2gnumber_olds
        site2gnumber_olds[cell_index]= value
        return_value = ','.join(site2gnumber_olds)
        return return_value 
    
    def value_for_import_ghi_chu(self,value):
        return_value = super(Excel_moto_onair,self).value_for_import_ghi_chu(value,'2G: ',insert_index=1)
        return return_value
class Excel_huawei_onair(Excel_2_3g):
    #for_purpose_thong_bao = False
    begin_row=0
    offset_begin_data_row = 0
    import_ghi_chu_text = 'CONFIG - ONAIR 2G HUAWEI W40.xlsx'
    using_function_if_empty_fields_or_field_not_in_excel_field = ['Site_ID_2G','Site_type','import_ghi_chu','active_2G']
    backwards_sequence =['brand_2G','BSC_2G','nha_san_xuat_2G']
    auto_map = False
    just_create_map_field = False
    update_or_create_main_item = 'Site_Name_1'
    worksheet_name = u'CONFIG'
    mapping_function_to_value_dict = {
                                      }
    manual_mapping_dict = {'Site_Name_1':u'Cell name','nha_san_xuat_2G':u'Loai Thiet Bi',\
                           'BSC_2G':'MBSC','Cell_ID_2G':'CI','Site_ID_2G_Number':'BTS Name','brand_2G':'Vendor'
                }
    def value_for_Site_Name_1 (self,value):
        self.cell_nth =  int(value[-5])
        if self.cell_nth ==0:
            self.cell_nth =1
        value = value[:-5] + value[-4:]
        print value
        return value
    def value_for_Site_ID_2G(self,cell_value):#6
        if self.created_or_update ==1:
            self.obj.nguoi_tao = User.objects.get(username='rnoc2')
            self.obj.ngay_gio_tao = timezone.now()
        prefix='HUA'
        cell_value = prefix.upper() + '_2G_' + self.obj.Site_Name_1
        return  cell_value
    def value_for_Site_ID_2G_Number(self,value):
        rs = re.findall('[BTS|IBS]_(\d+)',value)
        rt = rs[0]
        print '8***',rt
        return rt
    def value_for_Cell_ID_2G(self,value):
        value = u'%0.f'%value
        print 'self.cell_nth',self.cell_nth
        cell_index = self.cell_nth - 1
        site2gnumber_old_text= getattr(self.obj,'Cell_ID_2G',None)
        if site2gnumber_old_text:
            site2gnumber_olds = site2gnumber_old_text.split(',')
        else:
            site2gnumber_olds = []
        for i in range(0,self.cell_nth):
            value_i = value if i ==cell_index else 'None'
            try:
                site2gnumber_olds[i]
            except IndexError:
                site2gnumber_olds.append(value_i)
        print value_i
        print 'site2gnumber_olds',site2gnumber_olds
        site2gnumber_olds[cell_index]= value
        print 'site2gnumber new',site2gnumber_olds
        return_value = ','.join(site2gnumber_olds)
        return return_value 
    
    def value_for_import_ghi_chu(self,value):
        return_value = super(Excel_huawei_onair,self).value_for_import_ghi_chu(value,'2G: ',insert_index=1)
        return return_value
class Excel_ung_cuu_pro(Excel_2_3g):
    first_calculate_field = ['type_2g_or_3g']
    many2manyFields = ['comments']
    model = Mll
    begin_row=2
    offset_begin_data_row = 0
    import_ghi_chu_text = 'CONFIG - ONAIR 2G SRAN W40.xlsx'
    using_function_if_empty_fields_or_field_not_in_excel_field = ['ung_cuu','thiet_bi','brand','site_name','nguoi_tao','trang_thai','comments','ngay_gio_comment_cuoi_cung']
    backwards_sequence =['ngay_gio_comment_cuoi_cung','site_name','comments']
    auto_map = False
    just_create_map_field = False
    update_or_create_main_item = ['object','gio_mat']
    worksheet_name = u'Sheet1'
    mapping_function_to_value_dict = {
                                      }
    manual_mapping_dict = {'su_co':u'Loại sc','gio_tot':u'Ngày tốt','ngay_gio_tao':u'Ngày báo','comments':u'Ghi chú','object':u'ID','gio_mat':u'Ngày mll','type_2g_or_3g':u'Type','nguoi_tao':u'Người báo'}
    
    def value_for_ung_cuu(self,value):
        return True
    def value_for_trang_thai(self,value):
        tt = TrangThai.objects.get(Name=u'Báo ứng cứu')
        return tt
    def su_dung_obj_function(self):
        self.obj.cm.mll = self.obj
        self.obj.cm.save()
    def value_for_nguoi_tao(self, cell_value):
        print cell_value
        if cell_value==None:
            user = User.objects.get (username = 'rnoc2')
            return user
        else:
            user = UserProfile.objects.get (Name = cell_value).user
            return user
    def value_for_ngay_gio_tao(self,cell_value):#chinh
        return Excel_2_3g.value_for_ngay_gio_tao(self,cell_value,datetimeformat = '%d/%m/%Y %H:%M')
    def value_for_ngay_gio_comment_cuoi_cung(self,value):
        return self.obj.ngay_gio_tao
         
    def value_for_comments(self,value):
        print '*************',value
        if value ==None:
            value =''
        try:
            cm = Comment.objects.get(nguoi_tao=self.obj.nguoi_tao,trang_thai = TrangThai.objects.get(Name=u'Báo ứng cứu'),
                     datetime = self.obj.ngay_gio_tao,comment=value,ngay_gio_tao= self.obj.ngay_gio_tao )
            print '***********c o roi'
        except:
            cm = Comment(nguoi_tao=self.obj.nguoi_tao,trang_thai = TrangThai.objects.get(Name=u'Báo ứng cứu'),
                     datetime = self.obj.ngay_gio_tao,comment=value,ngay_gio_tao= self.obj.ngay_gio_tao )
        #cm.save()
        #self.obj.comments.add(cm)
        self.obj.cm =cm
        return None
        
    def value_for_gio_mat(self, value):
        d = local_a_naitive(datetime.datetime.strptime(value, '%d/%m/%Y %H:%M'))
        return d
    def value_for_gio_tot(self, value):
        d = local_a_naitive(datetime.datetime.strptime(value, '%d/%m/%Y %H:%M'))
        return d
    def value_for_type_2g_or_3g(self,value):
        
        if value == 'SRAN':
            value = '3G'
        if value !='3G' or value !='2G':
            return None
        return Excel_2_3g.value_for_Foreinkey(self,value,Class_pass_to=BTSType,is_tao = False)
    def value_for_su_co(self,value):
        return Excel_2_3g.value_for_Foreinkey(self,value,Class_pass_to=SuCo,is_tao = True,nguoi_tao = User.objects.get(username="rnoc2"))
    def value_for_object(self,value):
        #return value
        
        '''
        try:
            tramobj = Tram.objects.filter(Q(Site_Name_1=value)|Q(Site_ID_3G__icontains=value))[0]
        except IndexError:
            print 'chua co tram',value
            return value
        '''
        if self.temp_obj ['type_2g_or_3g'] == None:
            filter = Tram.objects.filter(Q(Site_Name_1=value))
            if len(filter) > 0:
                self.temp_obj ['type_2g_or_3g'] = Excel_2_3g.value_for_Foreinkey(self,'2G',Class_pass_to=BTSType,is_tao = False)
                tramobj = filter[0]
            else:
                filter = Tram.objects.filter(Q(Site_ID_3G__icontains=value))
                if len(filter) > 0:
                    self.temp_obj ['type_2g_or_3g'] = Excel_2_3g.value_for_Foreinkey(self,'3G',Class_pass_to=BTSType,is_tao = False)
                    tramobj = filter[0]
                else:
                    tramobj = None
        else:
            try:
                tramobj = Tram.objects.filter(Q(Site_Name_1=value)|Q(Site_ID_3G__icontains=value))[0]
            except IndexError:
                tramobj = None
        if tramobj==None:
            return value     

        if self.temp_obj ['type_2g_or_3g'].Name =='2G':
            object = tramobj.Site_ID_2G
            self.temp_obj['thiet_bi'] = tramobj.nha_san_xuat_2G
            self.temp_obj['brand'] = tramobj.nha_san_xuat_2G.brand if tramobj.nha_san_xuat_2G else None
            
        elif self.temp_obj ['type_2g_or_3g'].Name =='3G':
            object = tramobj.Site_ID_3G
            self.temp_obj['brand'] = tramobj.Cabinet.brand if tramobj.Cabinet else None
            self.temp_obj['thiet_bi'] = tramobj.Cabinet
        if object ==None:
            object = value
        return object  
        
        '''
        try:
            tramobj = Tram.objects.filter(Q(Site_Name_1=self.obj.object)|Q(Site_ID_3G=self.obj.object))[0]
            sitename = tramobj.Site_Name_1
            #self.obj.object = getattr(tramobj, key)
        except IndexError:
            
            sitename = None
        '''
    def value_for_site_name(self,value):
        if self.obj.type_2g_or_3g ==None:
            return None
        if self.obj.type_2g_or_3g.Name =='2G':
            key = 'Site_ID_2G'
        else:
            key = 'Site_ID_3G'
        karg = {'%s__iregex'%key:self.obj.object+'$'}
        try:
            tramobj = Tram.objects.filter(Q(**karg)|Q(Site_Name_1=self.obj.object))[0]
            sitename = tramobj.Site_Name_1
            #self.obj.object = getattr(tramobj, key)
        except IndexError:
            sitename = None
        return sitename
class Excel_sran_onair(Excel_2_3g):
    #for_purpose_thong_bao = False
    
    begin_row=0
    offset_begin_data_row = 0
    import_ghi_chu_text = 'CONFIG - ONAIR 2G SRAN W40.xlsx'
    using_function_if_empty_fields_or_field_not_in_excel_field = ['Site_ID_2G','Site_type','import_ghi_chu','active_2G','nguoi_tao','brand_2G','ngay_gio_sua']
    backwards_sequence =['brand_2G','nha_san_xuat_2G','BSC_2G',]
    auto_map = False
    just_create_map_field = False
    update_or_create_main_item = 'Site_Name_1'
    worksheet_name = u'Config_W40'
    mapping_function_to_value_dict = {'BSC_2G':'value_for_RNC'
                                      }
    manual_mapping_dict = {'Site_Name_1':u'CellName','nha_san_xuat_2G':u'Cabinet',\
                           'BSC_2G':u'MBSC','Cell_ID_2G':u'CI',
                }
    
    def value_for_brand_2G(self,value):
        value = 'Ericsson' 
        return super(Excel_sran_onair, self).value_for_brand_2G(value)
    def value_for_Site_Name_1 (self,value):
        self.cell_nth =  int(value[-5])
        
        if self.cell_nth ==0:
            self.cell_nth =1
        value = value[3:-5] + value[-4:]
        print value
        return value
    def value_for_Site_ID_2G(self,cell_value):#7
        prefix='SRN'
        cell_value = prefix.upper() + '_2G_' + self.obj.Site_Name_1
        return  cell_value
    def value_for_Site_ID_2G_Number(self,value):
        rs = re.findall('[BTS|IBS]_(\d+)',value)
        rt = rs[0]
        print '8***',rt
        return rt
    def value_for_Cell_ID_2G(self,value):
        value = u'%0.f'%value
        print 'self.cell_nth',self.cell_nth
        cell_index = self.cell_nth - 1
        site2gnumber_old_text= getattr(self.obj,'Cell_ID_2G',None)
        if site2gnumber_old_text:
            site2gnumber_olds = site2gnumber_old_text.split(',')
        else:
            site2gnumber_olds = []
        for i in range(0,self.cell_nth):
            value_i = value if i ==cell_index else 'None'
            try:
                site2gnumber_olds[i]
            except IndexError:
                site2gnumber_olds.append(value_i)
        print value_i
        print 'site2gnumber_olds',site2gnumber_olds
        site2gnumber_olds[cell_index]= value
        print 'site2gnumber new',site2gnumber_olds
        return_value = ','.join(site2gnumber_olds)
        return return_value 
    def value_for_import_ghi_chu(self,value):
        return_value = super(Excel_sran_onair,self).value_for_import_ghi_chu(value,'2G: ',insert_index=1)    
    

class Excel_3g_nsn_on_air(Excel_2_3g):
    #for_purpose_thong_bao = False
    begin_row=0
    offset_begin_data_row = 0
    import_ghi_chu_text = 'CONFIG - ONAIR 2G MOTO W40.xlsx'
    using_function_if_empty_fields_or_field_not_in_excel_field = ['Site_ID_3G','active_3G','Site_type','import_ghi_chu','nguoi_tao','brand_3G','Cabinet']
    backwards_sequence =['brand_3G','RNC','Cell_ID_3G','Site_ID_3G','Cabinet']
    auto_map = False
    just_create_map_field = False
    update_or_create_main_item = ['Site_Name_1']
    worksheet_name = u'Config'
    mapping_function_to_value_dict = {
                                      }
    manual_mapping_dict = {'Site_Name_1':u'Cell name','Cell_ID_3G':u'Cell ID','LAC_3G':'LAC','RNC':'MBSC Name'
                }
    def value_for_brand_3G(self,value):
        value = 'NSN'
        return super(Excel_3g_nsn_on_air, self).value_for_brand_2G(value)
    def value_for_Cabinet(self,cell_value):
        value = 'NSN Cabinet'
        return_value = Excel_2_3g.value_for_Cabinet(self,value)
        return  return_value
    
    def value_for_Site_Name_1 (self,cell_value):
        self.cell_nth =  int(cell_value[-5])
        sitename1 = cell_value[3:-6] + cell_value[-4:]
        self.site_id_3g = sitename1
        print sitename1
        return sitename1
    
    def value_for_LAC_3G(self,value):
        if isinstance(value, float):
            return u'%0.f'%value
        else:
            print '***',value
            return value
    def value_for_Site_ID_3G(self,value):
        return 'NSN_3G_' + self.site_id_3g 
    def value_for_Cell_ID_3G(self,value):
        if isinstance(value, float):
            value = u'%0.f'%value
        print 'self.cell_nth',self.cell_nth
        cell_index = self.cell_nth - 1
        site2gnumber_old_text= getattr(self.obj,'Cell_ID_3G',None)
        if site2gnumber_old_text:
            site2gnumber_olds = site2gnumber_old_text.split(',')
        else:
            site2gnumber_olds = []
        for i in range(0,self.cell_nth):
            value_i = value if i ==cell_index else 'None'
            try:
                site2gnumber_olds[i]
            except IndexError:
                site2gnumber_olds.append(value_i)
        print 'site2gnumber_olds',site2gnumber_olds
        site2gnumber_olds[cell_index]= value
        return_value = ','.join(site2gnumber_olds)
        return return_value 
    def value_for_import_ghi_chu(self,value):
        return_value = super(Excel_3g_nsn_on_air,self).value_for_import_ghi_chu(value,'3G: ',insert_index=0)
        return return_value
    
class Excel_3g_alu_onair(Excel_3g_nsn_on_air):
    worksheet_name = u'3G Config'
    import_ghi_chu_text = 'CONFIG - ONAIR 3G ALU W40.xlsx'  
    manual_mapping_dict = {'Site_Name_1':u'Cell name','Cell_ID_3G':u'Cell ID','LAC_3G':'LAC'}#bo di rnc
    def value_for_brand_3G(self,value):
        value ='Alcatel'
        return Excel_2_3g.value_for_brand_2G(self,value)
    def value_for_Cabinet(self,cell_value):
        cell_value = '9926 BBU'
        return_value = Excel_2_3g.value_for_Cabinet(self,cell_value)
        return  return_value
    def value_for_Site_Name_1 (self,cell_value):
        self.cell_nth =  int(cell_value[-5])
        sitename1 = cell_value[3:-5] + cell_value[-4:]
        self.site_id_3g = sitename1
        print sitename1
        return sitename1
    def value_for_Site_ID_3G(self,value):
        return 'ALU_3G_' + self.site_id_3g 

    
class Excel_3g_ericsson_onair(Excel_3g_nsn_on_air):# lay them duoc cell ID
    many2manyFields = ['du_an']
    offset_begin_data_row = 0
    worksheet_name = u'Config'
    import_ghi_chu_text = 'CONFIG - ONAIR 3G ERI W40.xlsx'   
    using_function_if_empty_fields_or_field_not_in_excel_field = ['nha_tram','Site_ID_3G','active_3G','Site_type','import_ghi_chu','nguoi_tao','brand_3G','ngay_gio_tao']
    manual_mapping_dict = {'Site_Name_1':u'Cell name',
                           'Cell_ID_3G':u'Cell ID',
                           'LAC_3G':'LAC',
                           'RNC':'MRNC Name',
                           'Site_ID_3G':u'siteid',
                           'tan_so_3G':u'Site Type'
                
                }
    
    def value_for_import_ghi_chu(self,value):
        return_value = Excel_2_3g.value_for_import_ghi_chu(self,value,'3G_onair: ',insert_index=3)
        return return_value
    def value_for_tan_so_3G(self,value):
        print 'tanso_________',value
        try:
            value = value.replace(' only','')
        except AttributeError: #'int' object has no attribute 'replace':
            return u'bỏ'
        return super(Excel_3g_ericsson_onair, self).value_for_Foreinkey(value,Class_pass_to=TanSo3G)
    def value_for_nha_tram(self,value):
        is_k = self.obj.Site_Name_1[-5]
        if is_k =='K':
            nha_tram_name = self.obj.Site_Name_1[0:-5] + self.obj.Site_Name_1[-4:]
            print '******nha_tram_name',nha_tram_name
            self.obj.is_tram_k=True
        else:
            nha_tram_name = self.obj.Site_Name_1
            self.obj.is_tram_k = False
        return super(Excel_3g_ericsson_onair, self).value_for_Foreinkey(nha_tram_name,Class_pass_to=NhaTram,
                                                                        nguoi_tao = User.objects.get(username='rnoc2'),
                                                                        ngay_gio_tao = timezone.now(),is_limit_so_luong_tao_foreign=False)    
    def su_dung_obj_function(self):
        if self.obj.is_tram_k:
            self.obj.nha_tram.tram_3g2 = self.obj
        else:
            self.obj.nha_tram.tram_3g1 = self.obj
        self.obj.nha_tram.save()
        print '***********su_dung_obj_function ok'
        
        
    def value_for_Site_Name_1 (self,cell_value):
        try:
            self.cell_nth =  int(cell_value[-5])
            is_K= cell_value[-6]
            if is_K =='K' and self.cell_nth > 6:
                is_K = 'E'
                print '@@@@@@@@skip'
                #cell_value[-6] = 'E'
                #self.skip = True
        except ValueError:
            self.cell_nth = 1
        
        sitename1 = cell_value[3:-6] +is_K  + cell_value[-4:]
        print '**sitename1',sitename1
        return sitename1
    def value_for_brand_3G(self,value):
        value = 'Ericsson'
        return super(Excel_3g_nsn_on_air, self).value_for_brand_2G(value)
    def value_for_Site_ID_3G(self,value):
        return 'ERI_3G_' + value
class Excel_3g_ericsson_onair_for_tan_so(Excel_3g_nsn_on_air):
    offset_begin_data_row = 0
    worksheet_name = u'Onair'
    import_ghi_chu_text = 'tab Onair CONFIG - ONAIR 3G ERI W40.xlsx'   
    using_function_if_empty_fields_or_field_not_in_excel_field = ['active_3G','Site_type','import_ghi_chu','nguoi_tao','brand_3G','ngay_gio_tao']
    manual_mapping_dict = {'Site_Name_1':u'Site Name','tan_so_3G':u'Site Type'}
    def value_for_Site_Name_1(self,value):
        value = value.replace('3G_','')
        return value
    def value_for_tan_so_3G(self,value):
        print 'tanso_________',value
        try:
            value = value.replace(' only','')
        except AttributeError: #'int' object has no attribute 'replace':
            return u'bỏ'
    def value_for_brand_3G(self,value):
        value = 'Ericsson'
        return super(Excel_3g_nsn_on_air, self).value_for_brand_2G(value)
class Excel_booster(Excel_2_3g):
    #for_purpose_thong_bao = False
    begin_row=1
    offset_begin_data_row = 0
    import_ghi_chu_text = 'booster.xlsx'
    using_function_if_empty_fields_or_field_not_in_excel_field = ['Site_type']
    backwards_sequence =[]
    auto_map = False
    just_create_map_field = False
    update_or_create_main_item = 'booster'
    worksheet_name = u'Sheet1'
    mapping_function_to_value_dict = {
                                      }
    manual_mapping_dict = {'booster':u'Name','Site_Name_1':u'Name'
                }
    def value_for_Site_Name_1(self,value):
        if self.created_or_update ==1:
            self.obj.nguoi_tao = User.objects.get(username='rnoc2')
            self.obj.ngay_gio_tao = timezone.now()
        return value + "_booster"
             
class Excel_to_3g_location (Excel_2_3g):
    using_function_if_empty_fields_or_field_not_in_excel_field = ['Site_type']
    allow_create_one_instance_if_not_exit = False
    auto_map = False
    just_create_map_field = False
    update_or_create_main_item = 'Site_ID_3G'
    worksheet_name = u'3G Site Location'
    mapping_function_to_value_dict ={}
    manual_mapping_dict = {'Site_ID_3G':u'Site ID','dia_chi_3G':u'Location','Long_3G':u'Long','Lat_3G':u'Lat','dia_chi_3G_khong_dau':'Location'}
    def value_for_Site_ID_3G(self,cell_value):
        value = 'ERI_3G_' + cell_value
        return value
    
        
    
class Excel_to_2g_config_SRAN (Excel_2_3g):
    using_function_if_empty_fields_or_field_not_in_excel_field = ['Site_type','is_tram_co_du_thong_tin_2g','nguoi_tao','ngay_gio_tao','active_2G']
    begin_row=37
    auto_map = False
    just_create_map_field = False
    update_or_create_main_item = 'Site_Name_1'
    worksheet_name = u'2G SRAN HCM Config'
    mapping_function_to_value_dict ={}
    manual_mapping_dict = {'Site_Name_1':u'RSITE','TG_Text':u'TG','TRX_DEF':u'TRX DEF'}
    def value_for_TG_Text(self,value):
        rs = re.findall('RXOTG-(\d+)',value)
        if len(rs)==1:
            self.obj.TG = rs[0]
        elif len(rs)>1:
            self.obj.TG = rs[0]
            self.obj.TG_1800 = rs[1]
        return value
    def value_for_Site_Name_1 (self,cell_value):
        cell_value = cell_value.replace('2G_','')
        return cell_value
class Excel_NSM(Excel_2_3g):
    using_function_if_empty_fields_or_field_not_in_excel_field = ['nguoi_tao','active_3G','brand_3G','Site_type','import_ghi_chu','is_tram_co_du_thong_tin_3g']
    begin_row=1
    backwards_sequence =['Cabinet','RNC']
    just_create_map_field = False
    auto_map = False
    update_or_create_main_item = 'Site_Name_1'
    worksheet_name = u'NSN Database'
    mapping_function_to_value_dict ={'Ngay_Phat_Song_3G':'value_for_common_datefield','IUB_VLAN_ID':'value_for_common_VLAN_ID'}
    manual_mapping_dict = {'Site_Name_1':u'3G Site Name','Site_ID_3G':u'3G Site Name','Cabinet':u'Type',\
                    'Ngay_Phat_Song_3G':u'Ngày PS U900','RNC':u'RNC name','IUB_VLAN_ID':u'VLAN ID','IUB_DEFAULT_ROUTER':u'GW IP ',\
                    'IUB_HOST_IP':u'IP','MUB_SUBNET_PREFIX':u'Network IP','MUB_DEFAULT_ROUTER':u'TRS IP',\
                    'ntpServerIpAddressPrimary':u'NTP Primary IP','ntpServerIpAddressSecondary':u'NTP Secondary  IP','tinh':u'Province'
                    }
    def value_for_brand_3G(self,value):
        value = 'NSN'
        return super(Excel_NSM, self).value_for_brand_2G(value)
    def value_for_Cabinet(self,cell_value):
        value = 'NSN Cabinet'
        return_value = super(Excel_NSM, self).value_for_Cabinet(value)
        return  return_value
    def value_for_Site_Name_1 (self,cell_value):
        cell_value = cell_value.replace('3G_','')
        return cell_value
    def value_for_Site_ID_3G(self,cell_value):
        cell_value = 'NSN_'+ cell_value
        return cell_value
    

class Excel_ALU_tuan(Excel_2_3g):
    using_function_if_empty_fields_or_field_not_in_excel_field = ['nguoi_tao','ngay_gio_tao','Cabinet','brand_3G','Site_type','import_ghi_chu']
    backwards_sequence =['brand_3G','Cabinet','RNC',]
    begin_row=0
    just_create_map_field = False
    auto_map = False
    update_or_create_main_item = 'Site_Name_1'
    worksheet_name = u'Sheet1'
    mapping_function_to_value_dict ={'Ngay_Phat_Song_3G':'value_for_common_datefield','IUB_VLAN_ID':'value_for_common_VLAN_ID','MUB_VLAN_ID':'value_for_common_VLAN_ID'}
    manual_mapping_dict = {'Site_Name_1':u'NodeB Name','Site_ID_3G':u'NodeB Name',\
                    'RNC':u'RNC Nam','IUB_DEFAULT_ROUTER':u'Iub Default GW NodeB','IUB_HOST_IP':u'Iub NodeB Ip add',
                    'IUB_VLAN_ID':u'Iub Vlan','IUB_SUBNET_PREFIX':u'Iub Default GW NodeB',
                    'MUB_VLAN_ID':u'Mub Vlan','MUB_SUBNET_PREFIX':u'Mub Default GW NodeB',\
                    'MUB_HOST_IP':u'Mub NodeB Ip add','MUB_DEFAULT_ROUTER':u'Mub Default GW NodeB',\
                    'tinh':'Tinh',
                   
                    }
    def value_for_brand_3G(self,value):
        value ='Alcatel'
        return super(Excel_ALU_tuan, self).value_for_brand_2G(value)
    def value_for_Cabinet(self,cell_value):
        cell_value = 'Alcatel Cabinet'
        return_value = super(Excel_ALU_tuan, self).value_for_Cabinet(cell_value)
        return  return_value
    
    def value_for_Site_Name_1 (self,cell_value):
        cell_value = cell_value.replace('3G_','')
        return cell_value
    def value_for_Site_ID_3G(self,cell_value):
        self.obj.active_3G = True
        cell_value = 'ALU_'+ cell_value
        return cell_value
    

class Excel_4G(Excel_2_3g):
    
    using_function_if_empty_fields_or_field_not_in_excel_field = ['Site_type','import_ghi_chu','active_4G','nguoi_tao','brand_4G']
    just_create_map_field = False
    auto_map = False
    update_or_create_main_item = 'Site_Name_1'
    worksheet_name = u'Ericsson 4G'
    mapping_function_to_value_dict ={'eNodeB_ID_DEC':'value_for_common_VLAN_ID'}
    manual_mapping_dict = {'eNodeB_Name':u'eNodeB_Name','Site_Name_1':u'eNodeB_Name','eNodeB_ID_DEC':u'eNodeB_ ID(DEC)','eNodeB_Type':u'eNodeB_Type',     }
    def value_for_tinh(self, value):#Excel_4G
        value = self.obj.Site_Name_1[-3:]
        return super(Excel_4G,self).value_for_tinh(value)
    def value_for_import_ghi_chu(self,value):
        return_value = super(Excel_4G,self).value_for_import_ghi_chu(value,'4G: ',insert_index=2)
        return return_value
    def value_for_Site_Name_1 (self,cell_value):
        cell_value = cell_value.replace('4G_','')
        return cell_value
    def value_for_brand_4G(self,value):
        value ='Ericsson'
        return super(Excel_4G, self).value_for_brand_2G(value)
    def value_for_eNodeB_Type(self,value):
        return_value = super(Excel_4G, self).value_for_Cabinet(value, '4G')
        return  return_value
class Excel_4G_Phu_Quoc(Excel_2_3g):
    begin_row=2
    using_function_if_empty_fields_or_field_not_in_excel_field = ['Site_type','import_ghi_chu','active_4G','nguoi_tao','brand_4G','eNodeB_Type']
    just_create_map_field = False
    auto_map = False
    update_or_create_main_item = 'Site_Name_1'
    worksheet_name = u'LNBTS'
    mapping_function_to_value_dict ={'IUB_VLAN_ID_4G':'value_for_common_VLAN_ID','MUB_VLAN_ID_4G':'value_for_common_VLAN_ID'}
    manual_mapping_dict = {'eNodeB_Name':u'eNB Name','Site_Name_1':u'eNB Name', 
                           'IUB_VLAN_ID_4G':6,'IUB_SUBNET_PREFIX_4G':7,'IUB_HOST_IP_4G':9,'IUB_DEFAULT_ROUTER_4G':12,
                           'MUB_VLAN_ID_4G':14,'MUB_SUBNET_PREFIX_4G':15,'MUB_HOST_IP_4G':17,'MUB_DEFAULT_ROUTER_4G':18}
    def value_for_tinh(self, value):#Excel_4G
        value = self.obj.Site_Name_1[-3:]
        return super(Excel_4G,self).value_for_tinh(value)
    def value_for_import_ghi_chu(self,value):
        return_value = super(Excel_4G_Phu_Quoc,self).value_for_import_ghi_chu(value,'4G: ',insert_index=2)
        return return_value
    def value_for_Site_Name_1 (self,cell_value):
        cell_value = cell_value.replace('4G_','')
        return cell_value
    def value_for_brand_4G(self,value):
        value ='NSN'
        return super(Excel_4G_Phu_Quoc, self).value_for_brand_2G(value)
    def value_for_eNodeB_Type(self,value):
        return_value = super(Excel_4G_Phu_Quoc, self).value_for_Cabinet('NSN 4G Cabinet', '4G')
        return  return_value  
from django.contrib.auth.models import User, Group
from django.contrib.auth.models import Permission
from django.contrib.contenttypes.models import ContentType
def create_user():
    workbook = xlrd.open_workbook(MEDIA_ROOT+ '/document/DanhSachEmail.xls')
    worksheet = workbook.sheet_by_name(u'Sheet3')
    num_rows = worksheet.nrows - 1
    row_index = -1
    print 'num_rows@@@@@@',num_rows
    while row_index < num_rows:
        row_index += 1
        username =   read_excel_cell(worksheet, row_index, 6)
        sdt  =   read_excel_cell(worksheet, row_index, 5)
        groupname =   read_excel_cell(worksheet, row_index, 7)
        name =   read_excel_cell(worksheet, row_index, 0) 
        ex = User.objects.get_or_create (username = username)
        user = ex[0]
        user.set_password(username)
        user.save()                          
        group = Group.objects.get_or_create (name = groupname)[0]
        group.user_set.add(user)
        ex = UserProfile.objects.get_or_create(user =user)
        profile = ex[0]
        profile.so_dien_thoai=sdt
        #profile.color_code = "#%06x" % random.randint(0, 0xFFFFFF)
        profile.color_code = "black"
        ca_truc = CaTruc.objects.latest('id')
        profile.ca_truc = ca_truc
        profile.Name = name
        profile.save()
    more_users = ['ductu','rnoc2']
    for username in more_users:
        ex = User.objects.get_or_create (username = username)
        user = ex[0]
        if ex[1]:#Neu user = New                                
            user.set_password(username)
            user.save()  
        ex = UserProfile.objects.get_or_create(user =user)
        profile = ex[0]
        #profile.color_code = "#%06x" % random.randint(0, 0xFFFFFF)
        profile.color_code = "black"
        print 'change color'
        ca_truc = CaTruc.objects.latest('id')
        profile.ca_truc = ca_truc
        profile.save()
 
def create_diaban():
    data = {u'Miền Đông':{'ky_hieu':'DNB'},u'Bắc Sông Hậu':{'ky_hieu':u'Bắc Sông Hậu'}\
            ,u'Nam Sông Hậu':{'ky_hieu':u'Nam Sông Hậu'},u'Hồ Chí Minh':{'ky_hieu':u'HCM'}
            }
    for k,v in data.iteritems():
        look_dict = {'Name':k}
        update_dict = {'Name':k,'ky_hieu':v['ky_hieu'],"Name_khong_dau":unidecode(k)}
        try:
            instance = DiaBan.objects.get(**look_dict)
            print 'da co',update_dict
        except DiaBan.DoesNotExist:
            instance = DiaBan(**update_dict)
            instance.save()
            print 'da save',update_dict
            
            
def grant_permission_to_group():
    content_type = ContentType.objects.get_for_model(Mll)
    name_and_codes = [('d4_create_truc_ca_permission','Can truc ca'),('can add on modal code','can add on modal')]
    truc_ca_group = Group.objects.get_or_create (name = 'truc_ca')[0]
    for x in name_and_codes:
        permission = Permission.objects.get_or_create(codename=x[0],
                                           name=x[1],
                                           content_type=content_type)[0]
    
        truc_ca_group.permissions.add(permission)
def grant_permission_admin():
    content_type = ContentType.objects.get_for_model(Mll)
    code_and_names= [('d4_admin','Can admin')]
    user = User.objects.get (
                                            username = 'rnoc2',
                                            )
    for x in code_and_names:
        permission = Permission.objects.get_or_create(codename=x[0],
                                           name=x[1],
                                           content_type=content_type)[0]
    
        user.user_permissions.add(permission)
        
def check_permission_of_group():
    for username in ['rnoc2','lucvk']:
        user = User.objects.get_or_create (
                                            username = username,
                                            )[0]
        permission = Permission.objects.get(codename='d4_create_truc_ca_permission')
        #print 'username,user.has_perm',username,user.has_perm('drivingtest.d4_create_truc_ca_permission')


            
              
from django.template import Context,Template 


def tao_script(instance_site,ntpServerIpAddressPrimary = '',ntpServerIpAddressSecondary = '',\
                         ntpServerIpAddress1="",ntpServerIpAddress2=""):
    if (ntpServerIpAddressPrimary=='' or ntpServerIpAddress1==""):
        return None
    #print 'hello, wellcome to download'
    Cabinet = instance_site.Cabinet
    is_luu_o_cung_moi_file_output_rieng = False
    is_save_archive_to_disk= True
    #save_type = 'temporary_achive_output_script'#or ,khong xai, 
    #save_type = 'disk_achive_output_script'
    #save_type = 'save_to_disk_3_file'
    #chi de hieu rang achive object co the ghi len o cung hoac len file tam
    now = datetime.datetime.now()
    Site_ID_3G= instance_site.Site_ID_3G
    instance_site.now = now
    achive_path=None
    instance_site.ntpServerIpAddressPrimary = ntpServerIpAddressPrimary
    instance_site.ntpServerIpAddressSecondary = ntpServerIpAddressSecondary
    instance_site.ntpServerIpAddress1 = ntpServerIpAddress1
    instance_site.ntpServerIpAddress2 = ntpServerIpAddress2
    template_files =[]
    if "RBS6" in Cabinet.Name:
        type_rbs = "6000"
        path_template_directory = MEDIA_ROOT+ '/document/template_script/6000/'
    elif "RBS3" in Cabinet.Name:
        path_template_directory = MEDIA_ROOT+ '/document/template_script/3000/'
        type_rbs = "3000"
    for root, dirs, files in os.walk(path_template_directory):
        for file in files:
            template_files.append(file)
    for counts,template_file in enumerate(template_files):
        path_to_1_template_file =  path_template_directory + template_file
        template = read_file_from_disk (path_to_1_template_file)
        t = Template(template)
        c = Context({'site3g':instance_site})
        output = t.render(c)
        fname = Site_ID_3G + '_' + template_file
        folder_name = '5484692'
        new_directory_path = MEDIA_ROOT+ '/for_user_download_folder/' + folder_name + '/'
        if is_luu_o_cung_moi_file_output_rieng:
            filepath = new_directory_path  + fname
            save_file_to_disk(filepath,output,1)
        if counts==0:
            if is_save_archive_to_disk:
                achive_path = new_directory_path + Site_ID_3G +'.zip'#achive_path den o cung
            else:# dang dung
                achive_path = tempfile.TemporaryFile() # this time achive_path is template object file,achive_path la 1 object template
            archive_file = zipfile.ZipFile(achive_path, 'w', zipfile.ZIP_DEFLATED)# tao ra 1 object achive de write len path
        if is_luu_o_cung_moi_file_output_rieng:#luu o cung 3 file rieng re, de doi chieu voi download o giao dien web
            if not os.path.exists(new_directory_path):
                os.makedirs(new_directory_path)
            filepath = new_directory_path  + fname
            save_file_to_disk(filepath,output,1)
        archive_file.writestr(fname, output)#write to object theo kieu data voi file name la fname
    archive_file.close()
    print 'type(achive_path)',type(achive_path)        
    return achive_path, type_rbs,is_save_archive_to_disk # achive_path become tempt zip file

def create_ca_truc():
    for ca_truc_name in ['Moto','Alu','Huawei','Sran','Khác']:
        ex = CaTruc.objects.get_or_create(Name=ca_truc_name)
        instance = ex[0]
        if ex[1]:
            instance.is_duoc_tao_truoc = True
            #instance.nguoi_tao = User.objects.get(username = "rnoc2")
            #instance.ngay_tao = timezone.now()
            instance.save()
def create_type_site():
    for x in [u'Site 0 (RNC,BSC)',u'Site BTS/NodeB']:
        try:
            instance = SiteType.objects.get(Name = x)
        except SiteType.DoesNotExist:
            instance = SiteType(Name=x)
            instance.save()
def create_type_bts():
    for x in [u'2G',u'3G',u'4G',u'ALL Band']:
        try:
            instance = BTSType.objects.get(Name = x)
        except BTSType.DoesNotExist:
            instance = BTSType(Name=x)
            instance.save()
                        

def delete_edithistory_table3g():
    EditHistory.objects.filter(modal_name='Tram').delete()
from openpyxl import load_workbook

'''
def export_excel_keo_dai():
    response = HttpResponse()
    path_or_file_or_response = response
    wb = load_workbook(MEDIA_ROOT  + '/document/MLL_keo_dai.xlsx')
    response['Content-Disposition'] = 'attachment; filename="keodai.xlsx"'
    ws = wb['Tổng hợp ']
    ws.cell(row =23, column = 1).value = 2
    wb.save(path_or_file_or_response)
    return path_or_file_or_response
'''
def tao_bao_cao_tinh_hinh(querysets = None,yesterday_or_other = None,exclude_4 = None):
    print 'querysets len',len(querysets)
    select_day,return_for_bao_cao =  export_excel_bcn (querysets = querysets,yesterday_or_other = yesterday_or_other,
                                                                 only_return_querysets_for_bc_tinhhinh = True,exclude_4 = exclude_4)
    path_to_1_template_file = MEDIA_ROOT+ '/document/template_baocaotinhhinh/tinhhinh.txt'
    template = read_file_from_disk (path_to_1_template_file)
    t = Template(template)
    try:
        return_for_bao_cao ['tong_thoi_gian_all_2G_3G'] = return_for_bao_cao ['tong_thoi_gian_all_2G'] + return_for_bao_cao ['tong_thoi_gian_all_3G']
    except KeyError:
        return 'khong co record mll nao'
    context = {'select_day':select_day,'count_2g':return_for_bao_cao['so_luong_mll_2G'],'tong_2g':return_for_bao_cao['tong_thoi_gian_all_2G']
               ,'count_3g':return_for_bao_cao['so_luong_mll_3G'],'tong_3g':return_for_bao_cao['tong_thoi_gian_all_3G'],
               'return_for_bao_cao':return_for_bao_cao
               }
    c = Context(context)
    output = t.render(c)
    return output
def xu_ly_bcn_1_row(bcn_record_row,min_select_time,max_select_time,
                    only_return_querysets_for_bc_tinhhinh,ws,begin_row,count,sheet,
                    one_more_loop_this_record_because_tach=0):
    if bcn_record_row.gio_mat < min_select_time :
        gio_mat = min_select_time
        
    else:
        gio_mat = bcn_record_row.gio_mat
    
    report_row_day = awaredate_time_to_local(gio_mat).date()#lay cai report_row_day
    if one_more_loop_this_record_because_tach > 0:
        report_row_day +=datetime.timedelta(days = one_more_loop_this_record_because_tach)
        gio_mat = local_a_naitive(datetime.datetime.combine(report_row_day,datetime.time(0,0,0)))
    if bcn_record_row.gio_tot:
        gio_tot = awaredate_time_to_local(bcn_record_row.gio_tot)
        if gio_tot > max_select_time:
            gio_tot = max_select_time
    else:#row_max_time neu khong co gio tot
        gio_tot = max_select_time
    cleaned_day  = awaredate_time_to_local(gio_tot).date()
    if cleaned_day > report_row_day:
        gio_tot = local_a_naitive(datetime.datetime.combine(report_row_day,datetime.time(23,59,59)))
        one_more_loop_this_record_because_tach += 1
    elif cleaned_day == report_row_day:
        one_more_loop_this_record_because_tach=None
    tong_thoi_gian = int(round((gio_tot -gio_mat).total_seconds()/60.0))
    if tong_thoi_gian <10:
        return tong_thoi_gian,one_more_loop_this_record_because_tach 
    if not only_return_querysets_for_bc_tinhhinh:
        ws.cell(row = begin_row + count, column = 6).value = gio_tot.strftime(TIME_FORMAT_FOR_BCN)    
        ws.cell(row = begin_row + count, column = 7).value = tong_thoi_gian
        ws.cell(row = begin_row + count, column = 5).value = awaredate_time_to_local(gio_mat).time().strftime(TIME_FORMAT_FOR_BCN)
        ws.cell(row = begin_row + count, column = 2).value = report_row_day.strftime(DATE_FORMAT_FOR_BCN)
        ws.cell(row = begin_row + count, column = 1).value = 2
        ws.cell(row = begin_row + count, column = 3).value = bcn_record_row.object
        ws.cell(row = begin_row + count, column = 4).value = bcn_record_row.BSC_or_RNC.Name
        ws.cell(row = begin_row + count, column = 8).value = str(bcn_record_row.code_loi)
        ws.cell(row = begin_row + count, column = 9).value = bcn_record_row.vnp_comment
        if sheet == "BCN_2G":
            ws.cell(row = begin_row + count, column = 10).value = bcn_record_row.gio_canh_bao_ac.strftime(DATETIME_FORMAT_FOR_BCN) if bcn_record_row.gio_canh_bao_ac else None
            ws.cell(row = begin_row + count, column = 11).value = bcn_record_row.object[-3:] 
        else:
            ws.cell(row = begin_row + count, column = 10).value = bcn_record_row.object[-3:]  
        return tong_thoi_gian,one_more_loop_this_record_because_tach    
    else:
        return tong_thoi_gian,one_more_loop_this_record_because_tach
    
def export_excel_bcn(querysets = None,yesterday_or_other = None,
                     is_for_keo_dai_download=False,
                     only_return_querysets_for_bc_tinhhinh = False,exclude_4 = False):
    if exclude_4:
        querysets = querysets.exclude(code_loi = 4)
    print 'yesterday_or_other****',yesterday_or_other
    if yesterday_or_other=='Today' or yesterday_or_other== 'Yesterday':
        if yesterday_or_other == 'Today':
            select_day =datetime.datetime.now()
        elif yesterday_or_other == 'Yesterday':
            select_day = datetime.datetime.now() - datetime.timedelta(days = 1)
        min_gio_mat_select_day = select_day
        min_select_time = local_a_naitive(datetime.datetime.combine(select_day,datetime.time(0,0,0)))
        to_day_for_filename = ''
        max_select_time = local_a_naitive(datetime.datetime.combine(select_day,datetime.time(23,59,59)))
        is_filter_query = False

    else:#theotable:
        is_filter_query = True
        min_gio_mat_select_day = awaredate_time_to_local(querysets.aggregate(Min('gio_mat'))['gio_mat__min']).date()
        min_select_time = local_a_naitive(datetime.datetime.combine(min_gio_mat_select_day,datetime.time(0,0,0)))
        max_gio_mat_select_day = awaredate_time_to_local(querysets.aggregate(Max('gio_mat'))['gio_mat__max']).date()
        max_select_time = local_a_naitive(datetime.datetime.combine(max_gio_mat_select_day,datetime.time(23,59,59)))
        to_day_for_filename = "-%s"%max_gio_mat_select_day.strftime('%d-%m-%Y')
    if is_for_keo_dai_download != True:
            q_group = Q(gio_mat__gte = min_select_time) & Q(gio_mat__lte = max_select_time)|\
            Q(gio_mat__lt = min_select_time)&Q(gio_tot__gte = min_select_time) \
           |Q(gio_mat__lt = min_select_time)&Q(gio_tot__isnull = True)
    else:
        print 'mll keo dai'
        delta_4_hours = max_select_time - datetime.timedelta(hours = 4)
        q_group = Q(gio_mat__lt = delta_4_hours)&Q(gio_tot__gte = max_select_time) \
           |Q(gio_mat__lt = delta_4_hours)&Q(gio_tot__isnull = True)
           
    if only_return_querysets_for_bc_tinhhinh:
        querysets =querysets.filter(q_group).exclude(code_loi = 7).exclude(code_loi = 8)#.exclude(code_loi = 4)
        return_for_bao_cao = {}
    querysets =querysets.filter(q_group).distinct()
    response = HttpResponse()
    path_or_file_or_response = response
    
    if is_for_keo_dai_download != True:
        response['Content-Disposition'] = 'attachment; filename="bcn_%s%s.xlsx"'%(min_gio_mat_select_day.strftime('%d-%m-%Y'),to_day_for_filename)
        wb = load_workbook(MEDIA_ROOT  + '/document/BCN_MLL_2G_3G_4G_02-05-2016.xlsx')
        SHEETS = ["BCN_2G","BCN_3G"]
        for sheet in SHEETS:
            if only_return_querysets_for_bc_tinhhinh:
                tong_tg_ma_loi_dict = {'ma_loi_5':0,'ma_loi_1':0}
                tong_thoi_gian_all = 0
            which_g = sheet.replace('BCN_','')

            ws = wb[sheet]
            max_row = ws.max_row
            if sheet == "BCN_2G":
                begin_row = 47
                range_xls  = 'A%s:J%s'%(str(begin_row),str(max_row))
                bts_type = '2G'
            elif sheet == "BCN_3G" :
                begin_row = 46
                range_xls  = 'A%s:I%s'%(str(begin_row),str(max_row))
                bts_type = '3G'
            for row in ws.iter_rows(range_xls):
                for cell in row:
                    cell.value = None
            mll_2g_or_3g_querysets = querysets.filter(BTS_Type__Name= bts_type).order_by('gio_mat').exclude(code_loi = 8).exclude(code_loi = 7)#.exclude(code_loi = 4)
            mll_2g_or_3g_querysets = mll_2g_or_3g_querysets.order_by('BTS_thiet_bi__Name')
            print 'len(bcns_2g)',len(mll_2g_or_3g_querysets)
            if len(mll_2g_or_3g_querysets) ==0:
                continue
            count = 0
            for bcn_record_row in mll_2g_or_3g_querysets:
                one_more_loop_this_record_because_tach=0
                while one_more_loop_this_record_because_tach !=None:
                    tong_thoi_gian,one_more_loop_this_record_because_tach = xu_ly_bcn_1_row(
                                                 bcn_record_row,min_select_time,max_select_time,
                                                  only_return_querysets_for_bc_tinhhinh,
                                                  ws,begin_row,count,sheet,
                                                  one_more_loop_this_record_because_tach)
                    if tong_thoi_gian< 10:
                        pass
                    else:
                        count +=1
                        if only_return_querysets_for_bc_tinhhinh:
                            key_ma_loi = 'ma_loi_%s'%bcn_record_row.code_loi 
                            if bcn_record_row.code_loi ==1 or bcn_record_row.code_loi==5:
                                tong_tg_ma_loi_dict [key_ma_loi] = tong_tg_ma_loi_dict.setdefault(key_ma_loi,0)  + tong_thoi_gian
                            tong_thoi_gian_all = tong_thoi_gian_all + tong_thoi_gian
            if only_return_querysets_for_bc_tinhhinh:
                return_for_bao_cao['so_luong_mll'+'_' + which_g ]   = count
                return_for_bao_cao['tong_thoi_gian_all'+'_' + which_g ]   = tong_thoi_gian_all
                for tong_thoi_gian_ma_loi in tong_tg_ma_loi_dict:
                    return_for_bao_cao['phan_tram_%s_'%tong_thoi_gian_ma_loi + which_g ]   = (100.0*tong_tg_ma_loi_dict[tong_thoi_gian_ma_loi]/tong_thoi_gian_all)
        if only_return_querysets_for_bc_tinhhinh:            
            return min_gio_mat_select_day.strftime(DATE_FORMAT_FOR_BCN),return_for_bao_cao
    
    
    
    
    else:#keo dai
        response['Content-Disposition'] = 'attachment; filename="bc_keo_dai_%s%s.xlsx"'%(min_gio_mat_select_day.strftime('%d-%m-%Y'),to_day_for_filename)
        wb = load_workbook(MEDIA_ROOT  + '/document/MLL_keo_dai.xlsx')
        ws = wb['Tổng hợp ']
        begin_row=22
        bcns_2g = querysets.filter(~Q(is_khong_tinh_mll_keo_dai = True)).order_by('gio_mat')
        for count,bcn_record_row in enumerate(bcns_2g):
            ws.cell(row = begin_row + count, column = 1).value = count + 1
            ws.cell(row = begin_row + count, column = 2).value = 2
            ws.cell(row = begin_row + count, column = 3).value = bcn_record_row.object
            #hethong
            ws.cell(row = begin_row + count, column = 4).value = bcn_record_row.BTS_Type.Name
            ws.cell(row = begin_row + count, column = 5).value = bcn_record_row.BTS_thiet_bi.brand.Name
            gio_mat = awaredate_time_to_local(bcn_record_row.gio_mat)
            ws.cell(row = begin_row + count, column = 6).value = gio_mat.strftime("%d/%m/%Y %H:%M ")
            ws.cell(row = begin_row + count, column = 7).value = bcn_record_row.object[-3:] 
            ws.cell(row = begin_row + count, column = 8).value = bcn_record_row.vnp_comment
            ws.cell(row = begin_row + count, column = 9).value = bcn_record_row.kien_nghi_de_xuat 
    wb.save(path_or_file_or_response)
    return path_or_file_or_response



def init_rnoc():
    
    create_ca_truc()#1
    create_user()#2
    import_database_4_cai_new(['ExcelImportTrangThai'])#3
    import_database_4_cai_new(['ExcelImportSuCo'])#4
    import_database_4_cai_new(['ExcelImportDuAn'])#5
    import_database_4_cai_new(['ExcelImportNguyenNhan'])#6
    #import_database_4_cai_new(['ExcelImportThietBi'])#7
    import_database_4_cai_new(['ExcelImportFaultLibrary'])#8
    import_database_4_cai_new(['ExcelImportThaoTacLienQuan'])#9
    #import_database_4_cai_new(['ExcelImportLenh'])#10
    import_database_4_cai_new(['ExcelImportDoiTac'])#11
    import_database_4_cai_new(['ExcelImportDoiTac_ungcuu'] )
    import_database_4_cai_new(['ImportTinh'] )
    create_diaban()
    import_database_4_cai_new(['ImportTinh_diaban'] )
    create_type_site()
    create_type_bts()
    import_database_4_cai_new(['Import_RNC_Tram'] )
from django.db.models.aggregates import Sum, Min, Count, Max
from django.db.models import Avg
from dateutil import rrule
def thong_ke_theo_ma_loi(qs,code_loi,tong_thoi_gian_mat,so_lan_mat_lien_lac):
        qs_loi = qs.filter(code_loi=code_loi)
        so_lan_mat_lien_lac_code = qs_loi.count()
        if so_lan_mat_lien_lac_code==0:
            so_lan_mat_lien_lac_percent = 0
            mat_dien_sum = 0
            avg_mat_dien = '_'
        else:
            so_lan_mat_lien_lac_percent = str(round(so_lan_mat_lien_lac_code*100/float(so_lan_mat_lien_lac),2))
            mat_dien_sum = qs_loi.aggregate(Sum('tong_thoi_gian'))['tong_thoi_gian__sum']
            avg_mat_dien =round(qs_loi.aggregate(Avg('tong_thoi_gian'))['tong_thoi_gian__avg'],2)
        ket_luan = mark_safe(u'{3} lần<span style="color:red">({4}%)</span>|{0} Phút({1}%)|{2} phút/lần'.format(mat_dien_sum,str(round(mat_dien_sum*100/float(tong_thoi_gian_mat),2)),avg_mat_dien,
                                                                         so_lan_mat_lien_lac_code,so_lan_mat_lien_lac_percent))
        return ket_luan

class ApiDataset(object):#thong ke bao cao ngay
    
    def __init__(self,BTS_type = 'all',MONTHLY_or_DAILY= 'MONTHLY',bg=None,end=None):
        if MONTHLY_or_DAILY== 'MONTHLY':
            bg = (datetime.datetime(2015, 12, 1, 0, 0, 0, 0))
            end =  (datetime.datetime(2016, 5, 12, 23, 59, 59, 79060))
        elif MONTHLY_or_DAILY== 'DAILY':
            #end = datetime.datetime.now()
            end = datetime.datetime(2016, 5, 12, 23, 59, 59, 79060)
            bg  = end -datetime.timedelta(days=14)
        self.ls = rrule.rrule(getattr(rrule,MONTHLY_or_DAILY), dtstart=bg, until=end)
        self.BTS_type = BTS_type
        self.MONTHLY_or_DAILY = MONTHLY_or_DAILY
        print 'self.lsself.lsself.lsself.lsself.lslen(self.ls)'
    '''
    def cache_data(self):
        # Access API and cache returned data on object.
        if self.data is None:
            self.data = 1
    '''
    def __iter__(self):
        so_lan_mat_lien_lac_prev = None
        for x in self.ls:
            if self.MONTHLY_or_DAILY== 'MONTHLY':
                thang_nam = x.strftime('%m/%Y')
            elif self.MONTHLY_or_DAILY== 'DAILY':
                thang_nam = x.strftime('%d/%m/%Y')
            karg = {'gio_mat__month':x.month,'gio_mat__year':x.year}
            if self.BTS_type =='all':
                pass
            else:
                karg.update({'BTS_Type__Name':self.BTS_type})
            if self.MONTHLY_or_DAILY== 'DAILY':
                karg.update({'gio_mat__day':x.day})
            qs = BCNOSS.objects.filter(**karg)
            so_lan_mat_lien_lac = qs.count()
            if not so_lan_mat_lien_lac_prev:
                so_lan_mat_lien_lac_Increase_or_descrease = u'_'
            else:
                so_lan_mat_lien_lac_Increase_or_descrease = ((so_lan_mat_lien_lac - so_lan_mat_lien_lac_prev )/float(so_lan_mat_lien_lac_prev))*100
            if isinstance(so_lan_mat_lien_lac_Increase_or_descrease, float):
                if so_lan_mat_lien_lac_Increase_or_descrease > 0 :
                    so_lan_mat_lien_lac_Increase_or_descrease = u'{0}%'.format(u'<span class="glyphicon glyphicon-arrow-up" style="color:red"></span>+%.1f'%so_lan_mat_lien_lac_Increase_or_descrease)
                else:
                    so_lan_mat_lien_lac_Increase_or_descrease = u'{0}%'.format(u'%.1f'%so_lan_mat_lien_lac_Increase_or_descrease)
            so_lan_mat_lien_lac_prev = so_lan_mat_lien_lac
            tong_thoi_gian_mat = qs.aggregate(Sum('tong_thoi_gian'))['tong_thoi_gian__sum']
            if so_lan_mat_lien_lac:
                thoi_gian_trung_binh_1_lan_mat ="{0:.2f}".format( round(qs.aggregate(Avg('tong_thoi_gian'))['tong_thoi_gian__avg'],1))
            else:
                thoi_gian_trung_binh_1_lan_mat = 0
                
            if self.BTS_type == '2G':
                tong_so_luong_tram_2g = Tram.objects.filter(active_2G=True).count()
            elif self.BTS_type == '3G':
                tong_so_luong_tram_2g = Tram.objects.filter(active_3G=True).count()
            else:
                tong_so_luong_tram_2g = Tram.objects.filter(Q(active_3G=True)).count()
            try :
                thoi_luong_mat_trung_binh_cua_1_tram_trong_thang = u"%.1f"%(tong_thoi_gian_mat/float(tong_so_luong_tram_2g))
                #thoi_luong_mat_trung_binh_cua_1_tram_trong_thang = tong_so_luong_tram_2g
            except ZeroDivisionError:
                thoi_luong_mat_trung_binh_cua_1_tram_trong_thang =u'_'
            except TypeError:
                thoi_luong_mat_trung_binh_cua_1_tram_trong_thang =u'_'

            if so_lan_mat_lien_lac==0:
                mat_dien_tong = u'_'
                truyen_dan_tinh_tong = u'_'
                thiet_bi_tong = u'_'
            else:
                mat_dien_tong = thong_ke_theo_ma_loi(qs,1,tong_thoi_gian_mat,so_lan_mat_lien_lac)
                truyen_dan_tinh_tong = thong_ke_theo_ma_loi(qs,5,tong_thoi_gian_mat,so_lan_mat_lien_lac)
                thiet_bi_tong = thong_ke_theo_ma_loi(qs,3,tong_thoi_gian_mat,so_lan_mat_lien_lac)
            so_lan_mat_lien_lac_txt = u'{0} Lần {1}'.format(so_lan_mat_lien_lac,so_lan_mat_lien_lac_Increase_or_descrease)
            data_item = {'thang_nam':thang_nam,'so_lan_mat_lien_lac':so_lan_mat_lien_lac_txt,\
                         'tong_thoi_gian_mat':u'%s phút'%tong_thoi_gian_mat,'thoi_gian_trung_binh_1_lan_mat':u'%s phút'%thoi_gian_trung_binh_1_lan_mat,\
                         'thoi_luong_mat_trung_binh_cua_1_tram_trong_thang':u'%s phút(tổng số trạm %s:%s)'%(thoi_luong_mat_trung_binh_cua_1_tram_trong_thang,self.BTS_type,tong_so_luong_tram_2g),\
                         'mat_dien_tong':mat_dien_tong,'truyen_dan_tinh_tong':truyen_dan_tinh_tong,'thiet_bi_tong':thiet_bi_tong,
                         }
            yield data_item 

    def __len__(self):
        return len(self.ls)


def thongkebcn_generator():
    
    bg = datetime.datetime(2014, 1, 1, 15, 29, 43, 79060)
    end = datetime.datetime(2016, 5, 12, 15, 29, 43, 79060)
    ls = rrule.rrule(rrule.MONTHLY, dtstart=bg, until=end)
    data_item = {}
    for x in ls:
        thang_nam = x.strftime('%m/%Y')
        qs = BCNOSS.objects.filter(gio_mat__month=x.month,gio_mat__year=x.year).exclude(code_loi = 8)
        so_lan_mat_lien_lac = qs.count()
        tong_thoi_gian_mat = qs.aggregate(Sum('tong_thoi_gian'))['tong_thoi_gian__sum']
        try:
            thoi_gian_trung_binh_1_lan_mat ="{0:.2f}".format( round(qs.aggregate(Avg('tong_thoi_gian'))['tong_thoi_gian__avg'],2))
        except TypeError:
            thoi_gian_trung_binh_1_lan_mat = None
        tong_so_luong_tram_2g = Tram.objects.filter(active_2G=True).count()
        try :
            thoi_luong_mat_trung_binh_cua_1_tram_trong_thang = u"%.2f"%(tong_thoi_gian_mat/float(tong_so_luong_tram_2g))
            #thoi_luong_mat_trung_binh_cua_1_tram_trong_thang = tong_so_luong_tram_2g
        except ZeroDivisionError:
            thoi_luong_mat_trung_binh_cua_1_tram_trong_thang =None
        except TypeError:
            thoi_luong_mat_trung_binh_cua_1_tram_trong_thang =None
        data_item = {'thang_nam':thang_nam,'so_lan_mat_lien_lac':so_lan_mat_lien_lac,\
                     'tong_thoi_gian_mat':tong_thoi_gian_mat,'thoi_gian_trung_binh_1_lan_mat':thoi_gian_trung_binh_1_lan_mat,'thoi_luong_mat_trung_binh_cua_1_tram_trong_thang':thoi_luong_mat_trung_binh_cua_1_tram_trong_thang}
        yield data_item    
        
def import_database_4_cai_new (runlists,workbook = None,import_ghi_chu = None):
        #'Excel_3G','Excel_to_2g',
        DB3G_SHEETS = ['Excel_3G','Excel_to_2g','Excel_to_2g_config_SRAN','Excel_to_3g_location','Excel_4G']
        is_db3g = False
        if 'ALL' in runlists:
            runlists.remove('ALL')
            runlists.extend(DB3G_SHEETS)
            runlists = unique_list(runlists)
        runlists_copy = runlists[:]
        runlists_reorder = []
        for x in DB3G_SHEETS:
            if x in runlists:
                is_db3g = True
                runlists_reorder.append(x)
                runlists_copy.remove(x)
        if is_db3g:
            for x in runlists_copy:
                runlists_reorder.append(x)
            runlists = runlists_reorder
            is_already_read_db3g_file  = False
        if workbook:
            for class_func_name in runlists:
                running_class = eval(class_func_name)
                return running_class(workbook = workbook,import_ghi_chu=import_ghi_chu).thong_bao
        else:
            for class_func_name in runlists:
                if class_func_name in DB3G_SHEETS:
                    if not is_already_read_db3g_file:
                        #path = MEDIA_ROOT+ u'/document/Ericsson_Database_Ver_161 - fortest.xlsx'
                        #path = MEDIA_ROOT+ u'/document/Ericsson_Database_Ver_161 - fortest.xlsx'
                        path = MEDIA_ROOT+ u'/document/Ericsson_Database_Ver_178.xlsx'
                elif class_func_name =='ImportRNC':
                    path = MEDIA_ROOT+ '/document/rnc.xls'
                elif class_func_name =='Excel_2g_rnas':
                    path = MEDIA_ROOT+ '/document/BTS_2009_rnas.xlsx'
                elif class_func_name =='Excel_moto_onair':
                    path = MEDIA_ROOT+ '/document/CONFIG - ONAIR 2G MOTO W40.xlsx'
                elif class_func_name =='Excel_huawei_onair':
                    path = MEDIA_ROOT+ '/document/W40/CONFIG - ONAIR 2G Huawei W40.xls'
                elif class_func_name =='Excel_sran_onair':
                    path = MEDIA_ROOT+ '/document/W40/CONFIG - ONAIR 2G SRAN W40.xlsx'
                elif class_func_name =='Excel_3g_alu_onair':
                    path = MEDIA_ROOT+ '/document/W40/CONFIG - ONAIR 3G ALU W40.xlsx'
                elif class_func_name =='Excel_3g_ericsson_onair':
                    path = MEDIA_ROOT+ '/document/W40/CONFIG_ONAIR 3G ERICSSON W40.xlsx'
                elif class_func_name =='Excel_3g_ericsson_onair_for_tan_so':
                    path = MEDIA_ROOT+ '/document/W40/CONFIG_ONAIR 3G ERICSSON W40.xlsx'
                elif class_func_name =='Excel_ericsson_onair':
                    path = MEDIA_ROOT+ '/document/CONFIG - ONAIR 2G ERICSSON  W40.xlsx' 
                elif class_func_name =='Excel_3g_nsn_on_air':
                    path = MEDIA_ROOT+ '/document/W40/CONFIG - ONAIR 3G NSN W40.xlsx'  
                elif class_func_name =='Excel_booster':
                    path = MEDIA_ROOT+ '/document/daily-report-booster.xlsx'
                elif class_func_name =='Excel_4G_Phu_Quoc':
                    path = MEDIA_ROOT+ '/document/eNB_TRS Parameter Planning v1.4 30122015.xls'      
                elif class_func_name =='Excel_NSM':
                    path = MEDIA_ROOT+ '/document/NSN_Database_version_4.xlsx'
                elif class_func_name =='Excel_ALU':
                    path = MEDIA_ROOT+ '/document/Database_ALU lot 1-2 -3 den NGAY  5-8-2015 .xls'
                elif class_func_name =='Excel_ALU_tuan':
                    path = MEDIA_ROOT+ '/document/alu_tuan.xlsx'
                elif class_func_name =='ExcelImportDoiTac_ungcuu':
                    path = MEDIA_ROOT+ '/document/danh sach nv xuong quan ly dia ban.xls'
                elif class_func_name =='ImportTinh':
                    path = MEDIA_ROOT+ '/document/danh sach nv xuong quan ly dia ban.xls'
                elif class_func_name =='ImportTinh_diaban':
                    path = MEDIA_ROOT+ '/document/To Ung cuu_New.tu.xls'
                elif class_func_name =='Import_RNC_Tram':
                    path = MEDIA_ROOT+ '/document/Table_BSCRNC.xls'
                elif class_func_name =='Import_BSCRNC':
                    path = MEDIA_ROOT+ '/document/Table_BSCRNC.xls'
                elif class_func_name =='Excel_3G':
                    path = MEDIA_ROOT+ '/document/3g_test.xls'
                elif class_func_name =='Excel_to_2g':
                    path = MEDIA_ROOT+ '/document/2g_test.xls'
                elif class_func_name =='ImportBCN2G':
                    path  = 'C:\Users\Administrator\Downloads\ReportWarning3G161228084159.xls'
                elif class_func_name =='ImportKeoDai':
                    path = MEDIA_ROOT+ '/document/input_keo_dai.xlsx'
                elif class_func_name =='ImportBCN2G_SRAN':
                    path = '/home/ductu/Documents/Downloads/2sran.xls'
                elif class_func_name =='ImportBCN3G':
                    path = '/home/ductu/Documents/Downloads/3g1.xls'
                elif class_func_name =='ImportBCN3G_ALU':
                    path = '/home/ductu/Documents/Downloads/ALU.xls'
                elif class_func_name =='ImportBCN3G_NSM':
                    path = '/home/ductu/Documents/Downloads/NSM.xls'
                elif class_func_name =='ExcelChung':
                    path = '/home/ductu/Documents/Downloads/Table_Tram.xls'
                elif class_func_name =='Excel_ung_cuu_pro':
                    #path = MEDIA_ROOT+ '/document/ucpro ton tai.xls'
                    path = MEDIA_ROOT+ '/document/ucpro da.xls'
                else: 
                    rs = re.match('^ExcelImport(.*?)$',class_func_name)
                    try:
                        classname = rs.group(1)
                        path = MEDIA_ROOT+ '/document/Table_%s.xls'%classname
                    except AttributeError:
                        raise ValueError('khong ton tai file name nao nhu the trong thu muc media')
                if class_func_name in DB3G_SHEETS:
                    if not is_already_read_db3g_file:     
                        workbook= xlrd.open_workbook(path)
                        is_already_read_db3g_file = True
                else:
                    workbook= xlrd.open_workbook(path)
                running_class = eval(class_func_name)
                import_ghi_chu = path[path.rfind('/')+1:]
                thong_bao = running_class(workbook = workbook,import_ghi_chu=import_ghi_chu).thong_bao
        return thong_bao
def so_tram_cho_tinh():
    for tinh_ins in Tinh.objects.all():
        print tinh_ins
        agg = Tram.objects.filter(tinh=tinh_ins).aggregate(count_3g = Count(F('Site_ID_3G')),count_2g = Count(F('Site_ID_2G')),\
                                                           count_all = Sum(Case(When(Q(active_3G = True)|Q(active_2G = True),then = 1)),output_field=IntegerField()))
        print agg
        tinh_ins.so_luong_tram_2G = agg['count_2g']
        tinh_ins.so_luong_tram_3G = agg['count_3g']
        tinh_ins.tong_so_tram = agg['count_all']
        tinh_ins.save()
        print 'save ok'
def so_tram_cho_RNC():
    for site0 in BSCRNC.objects.all():
        so_luong_tram = Tram.objects.filter(Q(RNC = site0)|Q(BSC_2G = site0)).count()
        if so_luong_tram ==0:
            so_luong_tram=1
        site0.so_luong_tram = so_luong_tram
        site0.save()
def set_password_for_user(str_username,new_pass):
    u = User.objects.get(username__exact=str_username)
    u.set_password(new_pass)
    u.save()
    print u'đã reset password thành công cho username %s'%str_username
if __name__ == '__main__':
    
    '''
    ist  =Tram (Site_Name_1='a',Site_type = SiteType.objects.get(Name=u'Site BTS/NodeB'),nguoi_tao= User.objects.get(username='tund'),ngay_gio_tao = timezone.now())
    ist.save()
    print 'dasave'
    
    '''
    #path = MEDIA_ROOT  + '\document\BCN_MLL_2G_3G_4G_02-05-2016.xlsx'
    #path = MEDIA_ROOT  + '\document\BCN20_11_2016.xlsx'
    #print path
    #wb = load_workbook(path)
    #select_day = timezone.now().date()
    #select_day = datetime.datetime.now()
    #print select_day
    #print timezone.now()
    #create_user()
    #import_database_4_cai_new(['ImportBCN2G'])
    #import_database_4_cai_new(['Excel_3G'])
    #import_database_4_cai_new(['Excel_ung_cuu_pro'])
    #import_database_4_cai_new(['ImportKeoDai'])
    #import_database_4_cai_new(['Excel_sran_onair'])
    #import_database_4_cai_new(['Excel_3g_ericsson_onair_for_tan_so'])
    #import_database_4_cai_new(['Excel_to_2g_config_SRAN'])
    #import_database_4_cai_new(['Excel_to_2g_config_SRAN'])
    '''
    instance = Tram.objects.filter(Site_Name_1 = 'test',Cabinet = ThietBi.objects.latest('id'),Site_Name_2 = None,Site_type  = SiteType.objects.get(Name = "Site BTS/NodeB"),nguoi_tao = User.objects.get(username='tund'))[0]
    print instance.Cabinet
    instance.Cabinet = None
    instance.save()
    print instance.Cabinet
    '''
    '''
    instance = Tram(Site_Name_1 = 'test',Cabinet = ThietBi.objects.latest('id'),Site_Name_2 = None,Site_type  = SiteType.objects.get(Name = "Site BTS/NodeB"),nguoi_tao = User.objects.get(username='tund'))
    instance.save()
    '''
    #create_type_site()
    #init_rnoc()
    #tb = import_database_4_cai_new(['ALL'])
    #tb = import_database_4_cai_new(['Excel_3g_alu_onair'])
    #tb = import_database_4_cai_new(['ALL'])
    #tb = import_database_4_cai_new(['Excel_4G_Phu_Quoc'])
    #tb = import_database_4_cai_new(['Excel_2g_rnas'])
    #tb = import_database_4_cai_new(['Excel_to_2g'])
    #tb = import_database_4_cai_new(['Excel_moto_onair'])
    #tb = import_database_4_cai_new(['Excel_huawei_onair'])
    #tb = import_database_4_cai_new(['Excel_NSM'])
    #tb = import_database_4_cai_new(['Excel_3g_nsn_on_air'])
    #tb = import_database_4_cai_new(['Excel_3g_ericsson_onair'])
    #tb = import_database_4_cai_new(['Excel_booster'])
    
    #print tb
    pass
    #set_password_for_user('tund','228787')
    #create_diaban()
    #user = User.objects.get(username = 'ductu').userprofile.ca_truc
    #print user
    
    #import_database_4_cai_new(['ImportTinh_diaban'] )
    '''
    so_tram_cho_tinh()
    so_tram_cho_RNC()
    '''
    #print Tram.objects.filter(Site_Name_1='CTRNC36')
    #import_database_4_cai_new(['Import_RNC_Tram'] )
    '''
    IS_GET_HTML = True
    if IS_GET_HTML:
        self_session = requests.session()
        
        self_session.headers = {'Accept' :   'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
    'Accept-Encoding':    'gzip, deflate',
    'Accept-Language'  :  'en-US,en;q=0.5',
    'Cache-Control'  :  'no-cache',
    'Connection'  :  'keep-alive',
    'Content-Type'  :  'application/x-www-form-urlencoded; charset=UTF-8',
    'Pragma'   : 'no-cache',
    'User-Agent'    :'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:35.0) Gecko/20100101 Firefox/35.0',
    'X-Requested-With'    :'XMLHttpRequest',}
        
        
        login_data = {
                'username': 'tund',
                'password': '2Vxe!mgst',
                }
        loginurl = 'http://el.vnptnet.vn/net/login/index.php'
        r = self_session.post(loginurl, data=login_data)
        login_ct = r.content 
        #print login_ct
        
        #view de thi
        r = self_session.post ('http://el.vnptnet.vn/net/mod/quiz/view.php?id=1') 
        html_de_lay_session_key = r.content 
        
        #with open(MEDIA_ROOT +'/thi_list.html', 'wb') as f:
                        #f.write(html_de_lay_session_key)
                      
        
        soup = BeautifulSoup(html_de_lay_session_key)
        class_entry_name = '.quizstartbuttondiv input[name=cmid]'
        soup2 = soup.select('.quizstartbuttondiv')[0]
        
        cmid = soup2.select('input[name=cmid]')
        cmid = cmid[0]['value']
        sesskey = soup2.select('input[name=sesskey]')
        sesskey = sesskey[0]['value']
        
        data = {'cmid':cmid,'sesskey':sesskey}
        print data
        thuc_hien_lai_de_thi = 'http://el.vnptnet.vn/net/mod/quiz/startattempt.php'
        r = self_session.post (thuc_hien_lai_de_thi,data=data) 
        #print r.content 
        print(r.url)
        
        
        with open(MEDIA_ROOT +'/thi_1_bai.html', 'wb') as f:
                        f.write(r.content)
        
                        
        

        
        html = r.content
    else:
        html =  read_file_from_disk (MEDIA_ROOT +'/thi_1_bai.html')
    SAVE_DB = False
    soup_lambaithi = BeautifulSoup(html)
    cauhoi_div = soup_lambaithi.select('.deferredfeedback ')
    tra_loi_dict = {}
    for cauhoi in cauhoi_div :
        new_cau_hoi_instance = TracNghiem()
        cau_hoi =  cauhoi.select('.qtext')[0].get_text()
        is_multiple_choice_c =  cauhoi.select('.prompt')[0].get_text()
        #print is_multiple_choice
        is_multiple_choice =  True if is_multiple_choice_c ==u"Chọn một hoặc nhiều đáp án:" else False  
        if is_multiple_choice:
            print 'ok',is_multiple_choice_c
        if SAVE_DB:
            try:
                cau_hoi_instance = TracNghiem.objects.get(cau_hoi= cau_hoi)
                #cau_hoi_instance.__dict__.update({is_multiple_choice:is_multiple_choice})
                cau_hoi_instance.is_multiple_choice = is_multiple_choice
                cau_hoi_instance.save()
            except:
                cau_hoi_instance = TracNghiem(cau_hoi= cau_hoi,is_multiple_choice=is_multiple_choice)
                cau_hoi_instance.save()
        print '***'
        if SAVE_DB:
            for traloi in cauhoi.select('.answer label'):
                traloi_content =  traloi.get_text()
                cau_hoi_instance.dap_an.get_or_create(Name = traloi_content)
        inputs =  cauhoi.select('.answer input[type=radio]')
        #inputs = cauhoi.find_all('input',type='checkbox',id=True)
        #print inputs
        #print len(inputs)
        if len(inputs) !=0:# radio name input are same
            tra_loi_dict.update({inputs[0]['name']:inputs[0]['value']})
        else:# checkbox
            inputs =  cauhoi.select('.answer input[type=checkbox]')
            tra_loi_dict.update({inputs[0]['name']:inputs[0]['value']})
        print len(inputs)    
        for input_ in inputs:
            print input_['name']
        
        for traloi in cauhoi.select('.answer label'):
                traloi_content =  traloi.get_text()
                print traloi_content
        
    print len(cauhoi_div)
    
    submit_url  ='http://el.vnptnet.vn/net/mod/quiz/processattempt.php'
    sesskey = soup_lambaithi.select('input[name=sesskey]')[0]['value']
    attempt = soup_lambaithi.select('input[name=attempt]')[0]['value']
    tra_loi_dict.update({'sesskey':sesskey,'attempt':attempt,'nextpage':'-1',
                         'slots':'1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24,25',
                         
                         })
    print tra_loi_dict
    r = self_session.post (submit_url,data=tra_loi_dict) 
    #print r.content 
    print(r.url)
    with open(MEDIA_ROOT +'/thi_1_bai_xong.html', 'wb') as f:
                        f.write(r.content)
                        
    '''